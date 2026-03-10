from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import UserChangeForm, UserCreationForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import get_user_model
from django.contrib.auth.models import BaseUserManager
from django.contrib.auth.hashers import make_password
from django.contrib.auth.tokens import default_token_generator
from django.contrib.sites.shortcuts import get_current_site
from django import forms
from django.db import transaction
from django.db.models import Sum, Count, F, Q, Prefetch
from django.forms import inlineformset_factory
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from django.utils.crypto import get_random_string
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.views import View
from django.views.decorators.http import require_POST, require_http_methods
from django.template.loader import render_to_string, get_template
from django.core.mail import EmailMessage, EmailMultiAlternatives
from django.views.decorators.csrf import csrf_exempt

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def detail_paiement(request, paiement_id):
    paiement = get_object_or_404(
        Paiement.objects.select_related('commande', 'commande__client'),
        id=paiement_id
    )
    return render(request, 'admin/detail_paiement.html', {'paiement': paiement})
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from collections import defaultdict
from datetime import timedelta
import json
import openpyxl
import pprint
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from weasyprint import HTML

from .models import (
    Produit, Commande, Paiement, Client, Agriculteur,
    Category, ReservationMateriel, CustomUser, LigneCommande,
    DetailCommande, Stock, MouvementStock, Notification, Materiel,
    Culture, Parcelle, ItineraireTechnique, ActiviteAgricole
)
from .forms import (
    ProduitForm, ClientForm, AgriculteurForm, MouvementStockForm,
    CommandeForm, DetailCommandeForm, ClientCreationForm, AddUtilisateurForm,
    CultureForm, ParcelleForm, ParcelleAssignationForm, ItineraireTechniqueForm,
    ActiviteAgricoleForm, ActiviteValidationForm
)
from .tokens import token_generator_24h
from .utils import envoyer_alerte_stock

User = get_user_model()

# Ici on crée le formset pour gérer les lignes de commande associées à une commande
DetailCommandeFormSet = inlineformset_factory(
    Commande, DetailCommande, form=DetailCommandeForm,
    fields=['produit', 'quantite'], extra=1, can_delete=True
)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def liste_utilisateurs(request):
    """
    Affiche la liste de tous les utilisateurs avec la possibilité de les activer/désactiver.
    Cette vue est accessible uniquement aux gerants.
    """
    # Récupérer tous les utilisateurs
    utilisateurs = User.objects.all().order_by('-date_joined')
    
    context = {
        'utilisateurs': utilisateurs,
        'titre': 'Liste des utilisateurs',
    }
    
    return render(request, 'admin/liste_utilisateurs.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def activer_utilisateur(request, user_id):
    """
    Active ou désactive un utilisateur en fonction de son état actuel.
    Cette fonction est utilisée pour basculer l'état d'activation d'un utilisateur.
    """
    user = get_object_or_404(User, id=user_id)
    user.is_active = not user.is_active
    user.save()
    
    status = "activé" if user.is_active else "désactivé"
    messages.success(request, f"L'utilisateur {user.username} a été {status}.")
    
    return redirect("liste_utilisateurs")

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def supprimer_utilisateur(request, user_id):
    """
    Supprime un utilisateur du système.
    Cette fonction est accessible uniquement aux gerants.
    """
    user = get_object_or_404(User, id=user_id)
    username = user.username
    
    try:
        user.delete()
        messages.success(request, f"L'utilisateur {username} a été supprimé avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression de l'utilisateur : {e}")
    
    return redirect("liste_utilisateurs")

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def modifier_utilisateur(request, user_id):
    """
    Affiche et traite le formulaire de modification d'un utilisateur.
    Cette fonction est accessible uniquement aux gerants.
    """
    user = get_object_or_404(User, id=user_id)
    
    if request.method == 'POST':
        form = AddUtilisateurForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, f"L'utilisateur {user.username} a été modifié avec succès.")
            return redirect('liste_utilisateurs')
    else:
        form = AddUtilisateurForm(instance=user)
    
    context = {
        'form': form,
        'user': user,
        'titre': f'Modifier l\'utilisateur {user.username}',
    }
    
    return render(request, 'admin/modifier_utilisateur.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def ajouter_utilisateur(request):
    """
    Affiche et traite le formulaire d'ajout d'un nouvel utilisateur.
    Cette fonction est accessible uniquement aux gerants.
    """
    if request.method == 'POST':
        form = AddUtilisateurForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.save()
            messages.success(request, f"L'utilisateur {user.username} a été ajouté avec succès.")
            return redirect('liste_utilisateurs')
    else:
        form = AddUtilisateurForm()
    
    context = {
        'form': form,
        'titre': 'Ajouter un nouvel utilisateur',
    }
    
    return render(request, 'admin/ajouter_utilisateur.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def activer_compte(request, uidb64, token):
    """
    Active le compte d'un utilisateur via un lien d'activation envoyé par email.
    """
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    
    if user is not None and token_generator_24h.check_token(user, token):
        user.is_active = True
        user.save()
        messages.success(request, "Le compte a été activé avec succès. L'utilisateur peut maintenant se connecter.")
    else:
        messages.error(request, "Le lien d'activation est invalide ou a expiré.")
    
    return redirect('liste_utilisateurs')

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def page_paiement(request):
    """
    Affiche la page de paiement pour effectuer un nouveau paiement.
    Cette vue est accessible uniquement aux administrateurs.
    """
    # Récupérer uniquement les commandes des clients (non staff)
    commandes = Commande.objects.filter(
        (Q(statut='en_attente') | Q(statut='en_cours')) &
        Q(client__is_staff=False)  # Exclure les commandes des gérants
    ).select_related('client').order_by('-date_commande')
    
    context = {
        'commandes': commandes,
        'title': 'Effectuer un paiement',
    }
    return render(request, 'paiement/page_paiement.html', context)

def initiate_payment(request):
    """
    Initialise un nouveau paiement.
    Cette vue est accessible uniquement aux administrateurs.
    """
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('login')
    
    if request.method == 'POST':
        try:
            client_id = request.POST.get('client')
            commande_id = request.POST.get('commande')
            montant = request.POST.get('montant')
            mode_paiement = request.POST.get('mode_paiement')
            
            client = get_object_or_404(Client, id=client_id)
            commande = get_object_or_404(Commande, id=commande_id, client=client)
            
            # Créer le paiement
            paiement = Paiement.objects.create(
                commande=commande,
                montant=montant,
                mode_paiement=mode_paiement,
                statut='valide',
                effectue_par=request.user
            )
            
            # Mettre à jour le statut de la commande
            commande.statut = 'payee'
            commande.save()
            
            # Créer une notification pour le client
            Notification.objects.create(
                user=client.user,
                message=f"Paiement de {montant} FCFA effectué avec succès pour la commande #{commande.id}.",
                type_notification='paiement_effectue'
            )
            
            messages.success(request, f"Paiement de {montant} FCFA enregistré avec succès pour la commande #{commande.id}.")
            return redirect('liste_commandes')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de l'enregistrement du paiement : {str(e)}")
    
    commande_prechargee = None
    client_precharge = None
    
    # Vérifier si une commande est spécifiée dans l'URL
    commande_id = request.GET.get('commande_id')
    if commande_id:
        try:
            commande_prechargee = Commande.objects.get(
                id=commande_id,
                statut='en_attente_paiement'
            )
            client_precharge = commande_prechargee.client
        except Commande.DoesNotExist:
            messages.warning(request, "La commande spécifiée n'existe pas ou n'est pas en attente de paiement.")
    
    # Récupérer la liste des clients actifs avec leurs commandes en attente de paiement
    clients = Client.objects.filter(est_actif=True).prefetch_related(
        Prefetch('commande_set', 
                queryset=Commande.objects.filter(statut='en_attente_paiement'),
                to_attr='commandes_en_attente'
        )
    )
    
    context = {
        'clients': clients,
        'commande_prechargee': commande_prechargee,
        'client_precharge': client_precharge,
        'title': 'Enregistrer un paiement',
        'modes_paiement': Paiement.MODE_PAIEMENT_CHOICES
    }
    return render(request, 'admin/initier_paiement.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def liste_paiements_client(request):
    """
    Affiche la liste des paiements et commandes validées regroupés par client.
    Cette vue récupère :
    1. Tous les paiements validés
    2. Toutes les commandes marquées comme validées
    Puis les regroupe par client avec le total des montants.
    """
    from django.db.models import Q
    import logging
    logger = logging.getLogger(__name__)
    
    # 1. Récupérer les paiements validés
    paiements_valides = (
        Paiement.objects
        .select_related('commande', 'commande__client')
        .filter(payment_status='Validated')
        .order_by('-payment_date')
    )
    
    # 2. Récupérer les commandes validées ou payées qui n'ont pas encore de paiement enregistré
    commandes_validees = Commande.objects.filter(
        Q(statut='validee') | Q(statut='payee')  # Inclure les commandes validées ou payées
    ).exclude(
        id__in=paiements_valides.values_list('commande_id', flat=True)
    ).select_related('client')
    
    # Log pour débogage
    logger.info(f"Requête SQL commandes validées: {str(commandes_validees.query)}")
    logger.info(f"Statuts de commande trouvés: {list(Commande.objects.values_list('statut', flat=True).distinct())}")
    
    # Ajout de logs pour le débogage
    logger.info(f"Requête commandes validées : {str(commandes_validees.query)}")
    logger.info(f"IDs des commandes avec paiements : {list(paiements_valides.values_list('commande_id', flat=True))}")
    
    logger.info(f"Paiements validés trouvés : {paiements_valides.count()}")
    logger.info(f"Commandes validées sans paiement : {commandes_validees.count()}")
    
    # 3. Créer un dictionnaire pour regrouper par client
    clients_data = {}
    
    # Log des données brutes pour le débogage
    logger.info("=== DONNÉES BRUTES POUR DÉBOGAGE ===")
    logger.info(f"Nombre de paiements validés : {paiements_valides.count()}")
    logger.info(f"Nombre de commandes validées : {commandes_validees.count()}")
    
    # Afficher les 5 premières commandes validées pour vérification
    for i, cmd in enumerate(commandes_validees[:5]):
        logger.info(f"Commande {i+1}: ID={cmd.id}, Statut='{cmd.get_statut_display()}', Client={cmd.client.username if cmd.client else 'None'}")
    
    # Afficher les 5 premiers paiements pour vérification
    for i, p in enumerate(paiements_valides[:5]):
        logger.info(f"Paiement {i+1}: ID={p.id}, Montant={p.amount}, Statut='{p.payment_status}', Commande ID={p.commande_id if p.commande else 'None'}")
    
    # Vérifier s'il y a des données à afficher
    if not paiements_valides.exists() and not commandes_validees.exists():
        logger.warning("Aucune donnée trouvée dans la base de données pour les critères de recherche")
        logger.info("Vérifiez les statuts dans la base de données avec la commande :")
        logger.info("from agricole.models import Commande; print(dict(Commande.STATUTS))")
    
    # 4. Traiter d'abord les paiements existants
    for paiement in paiements_valides:
        try:
            commande = paiement.commande
            client = commande.client
            
            if client.id not in clients_data:
                clients_data[client.id] = {
                    'name': client.get_full_name() or client.username,
                    'email': client.email,
                    'paiements': [],
                    'commandes': [],
                    'total': 0
                }
            
            # Ajouter le paiement
            payment_data = {
                'id': paiement.id,
                'type': 'paiement',
                'commande_id': commande.id,
                'montant': float(paiement.amount),
                'methode': paiement.get_payment_method_display(),
                'statut': paiement.get_payment_status_display(),
                'date': paiement.payment_date
            }
            clients_data[client.id]['paiements'].append(payment_data)
            clients_data[client.id]['total'] += float(paiement.amount)
            
        except Exception as e:
            logger.error(f"Erreur traitement paiement {paiement.id}: {str(e)}")
    
    # 5. Traiter les commandes validées sans paiement
    for commande in commandes_validees:
        try:
            client = commande.client
            
            if client.id not in clients_data:
                clients_data[client.id] = {
                    'name': client.get_full_name() or client.username,
                    'email': client.email,
                    'paiements': [],
                    'commandes': [],
                    'total': 0
                }
            
            # Ajouter la commande validée
            commande_data = {
                'id': commande.id,
                'type': 'commande',
                'montant': float(commande.get_total() or 0),
                'date': commande.date_commande,
                'statut': 'Validée',
                'details': f"Commande #{commande.id}"
            }
            clients_data[client.id]['commandes'].append(commande_data)
            clients_data[client.id]['total'] += float(commande.get_total() or 0)
            
        except Exception as e:
            logger.error(f"Erreur traitement commande {commande.id}: {str(e)}")
    
    # 6. Préparer les données pour le template
    clients_pour_template = {}
    for client_id, data in clients_data.items():
        # Fusionner paiements et commandes, trier par date
        tous_les_mouvements = data['paiements'] + data['commandes']
        tous_les_mouvements.sort(key=lambda x: x.get('date'), reverse=True)
        
        # Utiliser l'ID du client comme clé pour éviter les doublons de noms
        clients_pour_template[client_id] = {
            'id': client_id,
            'name': data['name'],
            'email': data['email'],
            'payments': tous_les_mouvements,
            'total': data['total'],
            'nb_paiements': len(data['paiements']),
            'nb_commandes': len(data['commandes'])
        }
    
    # 7. Trier les clients par montant total (décroissant)
    clients_tries = sorted(
        clients_pour_template.values(),
        key=lambda x: x['total'],
        reverse=True
    )
    
    context = {
        'clients_payments': {client['name']: client for client in clients_tries},
        'total_paiements': sum(c['total'] for c in clients_tries),
        'total_clients': len(clients_tries),
        'total_commandes_validees': sum(1 for c in clients_tries for p in c['payments'] if p['type'] == 'commande'),
        'total_paiements_valides': sum(1 for c in clients_tries for p in c['payments'] if p['type'] == 'paiement')
    }
    
    return render(request, 'admin/liste_paiements_client.html', context)

@login_required
def admin_home(request):
    # Vérifier que l'utilisateur est admin
    if not request.user.is_authenticated or request.user.role.lower() != 'admin':
        return redirect('login')
    
    # Comptes utilisateurs
    total_clients       = CustomUser.objects.filter(role__iexact='Client').count()
    total_agriculteurs  = CustomUser.objects.filter(role__iexact='Agriculteur').count()   # robustes

    # Produits / Commandes / Paiements
    total_produits           = Produit.objects.count()
    total_commandes          = Commande.objects.count()
    total_paiements          = Paiement.objects.count()
    total_paiements_valides  = Paiement.objects.filter(payment_status='validé').count()
    stock_total              = Produit.objects.aggregate(total_stock=Sum('quantite'))['total_stock'] or 0
    
    # Statistiques réservations matériel
    total_reservations = ReservationMateriel.objects.count()
    reservations_validees = ReservationMateriel.objects.filter(validee=True).count()
    reservations_en_attente = ReservationMateriel.objects.filter(validee=False).count()
    # Dernières réservations (5)
    dernieres_reservations = ReservationMateriel.objects.select_related('materiel', 'client').order_by('-date_reservation')[:5]

    # Activités des gerants pour le graphique circulaire
    admin_activities = {}
    
    # Récupérer tous les gerants
    admins = CustomUser.objects.filter(role__iexact='Admin')
    
    for admin in admins:
        admin_name = admin.get_full_name() or admin.username
        activities_count = 0
        
        # Compter les produits ajoutés par cet admin
        produits_count = Produit.objects.filter(fournisseur__isnull=True).count()
        activities_count += produits_count
        
        # Compter les commandes validées par cet admin (si applicable)
        commandes_validees = Commande.objects.filter(statut='validee').count()
        if admin.is_staff or admin.is_superuser:
            activities_count += commandes_validees // len(admins) if len(admins) > 0 else 0
        
        # Compter les paiements validés (si l'admin a validé)
        paiements_valides = Paiement.objects.filter(payment_status='Validated').count()
        if admin.is_staff or admin.is_superuser:
            activities_count += paiements_valides // len(admins) if len(admins) > 0 else 0
        
        # Définir la date limite avant de l'utiliser
        date_limite = datetime.now() - timedelta(days=30)  # ajuster selon tes besoins

        # Compter les utilisateurs créés par cet admin
        if admin.is_staff or admin.is_superuser:
            users_created = CustomUser.objects.filter(date_joined__gte=date_limite).count()
            activities_count += users_created // len(admins) if len(admins) > 0 else 0

        admin_activities[admin_name] = activities_count

    # Période dynamique
    periode = request.GET.get('periode', 'jour')
    now_dt  = timezone.now()

    deltas = {
        'jour':    timedelta(days=1),
        'semaine': timedelta(weeks=1),
        'mois':    timedelta(days=30),
        'annee':   timedelta(days=365),
    }
    date_limite = now_dt - deltas.get(periode, deltas['jour'])

    commandes_periode         = Commande.objects.filter(date_commande__gte=date_limite).count()
    paiements_valides_periode = Paiement.objects.filter(payment_status='validé',
                                                       payment_date__gte=date_limite).count()

    # Préparer les courbes sur 30 jours (commandes & paiements validés)
    start_date = (now_dt - timedelta(days=29)).date()
    dates = [start_date + timedelta(days=i) for i in range(30)]
    trend_labels = [d.strftime('%d/%m') for d in dates]

    # Compter par jour
    trend_commandes = [
        Commande.objects.filter(date_commande__date=d).count() for d in dates
    ]
    trend_paiements = [
        Paiement.objects.filter(payment_status__iexact='validé', payment_date__date=d).count() for d in dates
    ]

    # Activités récentes du gerant connecté (produits ajoutés)
    recent_products = (
        Produit.objects.filter(fournisseur=request.user)
        .order_by('-created_at')[:10]
    )
    recent_activities = []
    for p in recent_products:
        recent_activities.append({
            'type': 'stock',
            'title': f"Produit ajouté: {p.nom_produit}",
            'description': f"Type: {p.type_produit} • Quantité: {p.quantite} • Prix: {p.prix_unitaire}",
            'time': timezone.localtime(p.created_at).strftime('%d %b %Y %H:%M'),
        })

    context = {
        'total_clients': total_clients,
        'total_agriculteurs': total_agriculteurs,
        'total_produits': total_produits,
        'total_commandes': total_commandes,
        'total_paiements': total_paiements,
        'total_paiements_valides': total_paiements_valides,
        'stock_total': stock_total,
        'commandes_periode': commandes_periode,
        'paiements_valides_periode': paiements_valides_periode,
        'periode': periode,
        'now': now_dt,
        'total_reservations': total_reservations,
        'reservations_validees': reservations_validees,
        'reservations_en_attente': reservations_en_attente,
        'dernieres_reservations': dernieres_reservations,
        # Nouvelles données pour graphiques/activités
        'trend_labels': trend_labels,
        'trend_commandes': trend_commandes,
        'trend_paiements': trend_paiements,
        'recent_activities': recent_activities,
    }
    
    response = render(request, 'admin/admin_home.html', context)
    # Empêcher la mise en cache pour éviter l'accès via le bouton retour après déconnexion
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    response['X-Frame-Options'] = 'DENY'
    return response

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def materiels_agricoles(request):
    """
    Affiche la liste des matériels agricoles disponibles.
    Accessible uniquement aux gérants.
    """
    # Récupérer tous les matériels
    materiels = Materiel.objects.all().order_by('nom')
    
    # Statistiques
    total_materiels = materiels.count()
    materiels_disponibles = materiels.filter(disponible=True).count()
    materiels_indisponibles = materiels.filter(disponible=False).count()
    
    # Regrouper par catégorie
    categories = {}
    for materiel in materiels:
        categorie = materiel.categorie or 'Non catégorisé'
        if categorie not in categories:
            categories[categorie] = []
        categories[categorie].append(materiel)
    
    context = {
        'materiels': materiels,
        'categories': categories,
        'total_materiels': total_materiels,
        'materiels_disponibles': materiels_disponibles,
        'materiels_indisponibles': materiels_indisponibles,
        'total_reservations': ReservationMateriel.objects.count(),
        'reservations_en_cours': ReservationMateriel.objects.filter(validee=True).count()
    }
    
    return render(request, 'admin/materiels_agricoles.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def admin_reports(request):
    """
    Tableau de bord Rapports: KPIs, filtres période, et graphiques.
    Accessible uniquement aux gerants.
    """
    # Période sélectionnée
    periode = request.GET.get('periode', '30d')  # 7d, 30d, 90d, 365d
    mapping = {
        '7d': 7,
        '30d': 30,
        '90d': 90,
        '365d': 365,
    }
    days = mapping.get(periode, 30)

    now_dt = timezone.now()
    since = now_dt - timedelta(days=days)

    # KPIs basiques
    total_commandes = Commande.objects.count()
    total_paiements = Paiement.objects.count()
    total_produits = Produit.objects.count()
    total_clients = CustomUser.objects.filter(role__iexact='Client').count()

    # Activité sur la période
    commandes_period = Commande.objects.filter(date_commande__gte=since).count()
    paiements_valides_period = Paiement.objects.filter(payment_status__iexact='validé', payment_date__gte=since).count()

    # Top produits par ventes (approximatif via lignes de commande si dispo)
    top_produits = (
        DetailCommande.objects
        .values('produit__nom_produit')
        .annotate(qte=Sum('quantite'))
        .order_by('-qte')[:5]
    )

    context = {
        'periode': periode,
        'days': days,
        'total_commandes': total_commandes,
        'total_paiements': total_paiements,
        'total_produits': total_produits,
        'total_clients': total_clients,
        'commandes_period': commandes_period,
        'paiements_valides_period': paiements_valides_period,
        'top_produits': list(top_produits),
        'now': now_dt,
    }

    return render(request, 'admin/reports.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def admin_settings(request):
    """Page Paramètres gerant (profil, préférences, notifications, sécurité)."""
    context = {
        'user_obj': request.user,
        'last_login': getattr(request.user, 'last_login', None),
        'two_factor_enabled': False,  # Placeholder si 2FA n'est pas encore implémenté
        'email_notifications': True,  # Valeur par défaut
        'sms_notifications': False,   # Valeur par défaut
        'theme': 'light',             # light | dark
    }
    return render(request, 'admin/settings.html', context)
@login_required
def manage_users(request):
    # Récupérer tous les utilisateurs
    users = User.objects.all()

    return render(request, 'admin/manage_users.html', {
        'users': users,
    })

def add_user(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('manage_users')
    else:
        form = UserCreationForm()
    return render(request, 'admin/add_user.html', {'form': form})

def edit_user(request, id):
    user = User.objects.get(id=id)
    if request.method == 'POST':
        form = UserChangeForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            return redirect('manage_users')
    else:
        form = UserChangeForm(instance=user)
    return render(request, 'admin/edit_user.html', {'form': form})

def delete_user(request, id):
    user = get_object_or_404(User, id=id)
    user.delete()
    return redirect('manage_users')

def admin_profile(request):
    user = CustomUser.objects.get(id=request.user.id)

    context={
        "user": user
    }
    return render(request, 'admin/admin_profile.html', context)


def add_agriculteur(request):
    return render(request, "admin/add_agriculteur_template.html")


def add_agriculteur_save(request):
    if request.method != "POST":
        messages.error(request, "Méthode non autorisée.")
        return redirect('add_agriculteur')
    else:
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        adresse = request.POST.get('adresse')

        try:
            user = CustomUser.objects.create_user(
                username=username,
                password=password,
                email=email,
                first_name=first_name,
                last_name=last_name,
                user_type=2  # ou autre valeur correspondant aux agriculteurs
            )
            user.agriculteur.adresse = adresse
            user.save()
            messages.success(request, "Agriculteur ajouté avec succès !")
            return redirect('add_agriculteur')
        except Exception as e:
            messages.error(request, f"Erreur lors de l'ajout de l'agriculteur : {e}")
            return redirect('add_agriculteur')


def manage_agriculteurs(request):
    agriculteurs = Agriculteur.objects.all()
    context = {
        "agriculteurs": agriculteurs
    }
    return render(request, "admin/manage_agriculteur_template.html", context)


def edit_agriculteur(request, agriculteur_id):
    agriculteur = Agriculteur.objects.get(admin=agriculteur_id)
    context = {
        "agriculteur": agriculteur,
        "id": agriculteur_id
    }
    return render(request, "admin/edit_agriculteur_template.html", context)


def edit_agriculteur_save(request):
    if request.method != "POST":
        return HttpResponse("<h2>Méthode non autorisée</h2>")
    else:
        agriculteur_id = request.POST.get('agriculteur_id')
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        adresse = request.POST.get('adresse')

        try:
            user = CustomUser.objects.get(id=agriculteur_id)
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.username = username
            user.save()
            
            agriculteur_model = Agriculteur.objects.get(admin=agriculteur_id)
            agriculteur_model.adresse = adresse
            agriculteur_model.save()

            messages.success(request, "Agriculteur modifié avec succès.")
            return redirect('/edit_agriculteur/' + agriculteur_id)

        except Exception as e:
            messages.error(request, f"Erreur de mise à jour : {e}")
            return redirect('/edit_agriculteur/' + agriculteur_id)


def delete_agriculteur(request, agriculteur_id):
    agriculteur = Agriculteur.objects.get(admin=agriculteur_id)
    try:
        agriculteur.delete()
        messages.success(request, "Agriculteur supprimé avec succès.")
        return redirect('manage_agriculteurs')
    except Exception as e:
        messages.error(request, f"Erreur de suppression : {e}")
        return redirect('manage_agriculteurs')


def add_produit(request):
    if request.method == 'POST':
        form = ProduitForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Produit ajouté avec succès.")
            return redirect('manage_produit')
    else:
        form = ProduitForm()
    return render(request, "admin/add_product_template.html", {'form': form})





def manage_produit(request):
    products = Produit.objects.all()  # correspond à "products" dans le template
    context = {
        "products": products
    }
    return render(request, "admin/manage_produit.html", context)


def edit_product(request, produit_id):
    produit = get_object_or_404(Produit, id=produit_id)

    if request.method == 'POST':
        form = ProduitForm(request.POST, request.FILES, instance=produit)
        if form.is_valid():
            form.save()
            return redirect('manage_produit')  # ou le bon nom de ta URL
    else:
        form = ProduitForm(instance=produit)

    return render(request, 'admin/edit_product.html', {'form': form, 'produit': produit})


def delete_product(request, produit_id, force_delete=False):
    produit = get_object_or_404(Produit, id=produit_id)
    
    # Vérifier si le produit est utilisé dans des commandes
    from django.db.models import ProtectedError
    
    # Récupérer les commandes qui utilisent ce produit
    from .models import DetailCommande, LigneCommande
    
    # Vérifier les commandes qui utilisent ce produit
    details_commande = DetailCommande.objects.filter(produit=produit)
    lignes_commande = LigneCommande.objects.filter(produit=produit)
    
    # Si des commandes utilisent ce produit et que la suppression forcée n'est pas activée
    if (details_commande.exists() or lignes_commande.exists()) and not force_delete:
        # Préparer le contexte avec les informations sur les commandes
        context = {
            'produit': produit,
            'details_commande': details_commande,
            'lignes_commande': lignes_commande,
            'total_commandes': details_commande.count() + lignes_commande.count()
        }
        return render(request, 'admin/confirm_delete_product.html', context)
    
    # Si aucune commande ou suppression forcée demandée
    try:
        # D'abord supprimer les détails de commande liés
        details_commande.delete()
        lignes_commande.delete()
        
        # Puis supprimer le produit
        produit.delete()
        messages.success(request, "Produit supprimé avec succès.")
        
    except Exception as e:
        messages.error(request, f"Une erreur est survenue lors de la suppression : {str(e)}")
    
    return redirect('manage_produit')

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def delete_product_confirm(request, produit_id):
    """
    Vue pour confirmer et exécuter la suppression d'un produit, même s'il est utilisé dans des commandes.
    """
    if request.method == 'POST':
        # Appeler delete_product avec force_delete=True
        return delete_product(request, produit_id, force_delete=True)
    return redirect('manage_produit')

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def stock_movement(request):
    """
    Affiche l'historique des mouvements de stock avec possibilité de filtrage.
    Cette vue est accessible uniquement aux gerants.
    """
    # Récupérer les filtres depuis la requête
    produit_filtre = request.GET.get('produit')
    type_filtre = request.GET.get('type_mouvement')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    # Récupérer tous les mouvements de stock
    movements = MouvementStock.objects.all().order_by('-date')
    
    # Appliquer les filtres si nécessaire
    if produit_filtre:
        movements = movements.filter(produit_id=produit_filtre)
    
    if type_filtre:
        movements = movements.filter(type_mouvement=type_filtre)
    
    if date_debut:
        movements = movements.filter(date__gte=date_debut)
    
    if date_fin:
        movements = movements.filter(date__lte=date_fin)
    
    # Récupérer la liste des produits pour le filtre
    produits = Produit.objects.all().order_by('nom_produit')
    
    # Calculer les statistiques
    entrees = movements.filter(type_mouvement='ENTREE').count()
    sorties = movements.filter(type_mouvement='SORTIE').count()
    produits_concernes = movements.values('produit').distinct().count()
    
    context = {
        'movements': movements,
        'produits': produits,
        'entrees': entrees,
        'sorties': sorties,
        'produits_concernes': produits_concernes,
        'produit_filtre': produit_filtre,
        'type_filtre': type_filtre,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'titre': 'Historique des mouvements de stock',
    }
    
    return render(request, 'admin/stock_movements.html', context)
    
def add_client(request):
    form = ClientForm()
    context = {
        "form": form
    }
    return render(request, 'admin/add_client_template.html', context)
def add_client_save(request):
    if request.method != "POST":
        messages.error(request, "Méthode invalide")
        return redirect('add_client')
    else:
        form = ClientForm(request.POST, request.FILES)

        if form.is_valid():
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']
            username = form.cleaned_data['username']
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            address = form.cleaned_data['address']
            phone = form.cleaned_data['phone']

            if len(request.FILES) != 0:
                profile_pic = request.FILES['profile_pic']
                fs = FileSystemStorage()
                filename = fs.save(profile_pic.name, profile_pic)
                profile_pic_url = fs.url(filename)
            else:
                profile_pic_url = None

            try:
                user = CustomUser.objects.create_user(username=username,
                                                      password=password,
                                                      email=email,
                                                      first_name=first_name,
                                                      last_name=last_name,
                                                      user_type=4)  # Client type
                user.client.address = address
                user.client.phone = phone
                user.client.profile_pic = profile_pic_url
                user.save()
                messages.success(request, "Client ajouté avec succès !")
                return redirect('add_client')
            except Exception as e:
                messages.error(request, f"Échec de l'ajout du client : {e}")
                return redirect('add_client')
        else:
            return redirect('add_client')


def manage_client(request):
    clients = Client.objects.all()
    context = {
        "clients": clients
    }
    return render(request, 'admin/manage_client_template.html', context)


def edit_client(request, client_id):
    request.session['client_id'] = client_id
    client = Client.objects.get(admin=client_id)
    form = ClientForm(instance=client.admin)

    form.fields['email'].initial = client.admin.email
    form.fields['username'].initial = client.admin.username
    form.fields['first_name'].initial = client.admin.first_name
    form.fields['last_name'].initial = client.admin.last_name
    form.fields['address'].initial = client.address
    form.fields['phone'].initial = client.phone

    context = {
        "id": client_id,
        "form": form
    }
    return render(request, "admin/edit_client_template.html", context)


def edit_client_save(request):
    if request.method != "POST":
        return HttpResponse("Méthode invalide !")
    else:
        client_id = request.session.get('client_id')
        if client_id is None:
            return redirect('/manage_client')

        client = Client.objects.get(admin=client_id)
        form = ClientForm(request.POST, request.FILES, instance=client.admin)
        if form.is_valid():
            email = form.cleaned_data['email']
            username = form.cleaned_data['username']
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']
            address = form.cleaned_data['address']
            phone = form.cleaned_data['phone']

            if len(request.FILES) != 0:
                profile_pic = request.FILES['profile_pic']
                fs = FileSystemStorage()
                filename = fs.save(profile_pic.name, profile_pic)
                profile_pic_url = fs.url(filename)
            else:
                profile_pic_url = None

            try:
                user = CustomUser.objects.get(id=client_id)
                user.first_name = first_name
                user.last_name = last_name
                user.email = email
                user.username = username
                user.save()

                client_model = Client.objects.get(admin=client_id)
                client_model.address = address
                client_model.phone = phone
                if profile_pic_url:
                    client_model.profile_pic = profile_pic_url
                client_model.save()

                del request.session['client_id']
                messages.success(request, "Client mis à jour avec succès !")
                return redirect('/edit_client/' + str(client_id))
            except Exception as e:
                messages.error(request, f"Erreur lors de la mise à jour : {e}")
                return redirect('/edit_client/' + str(client_id))
        else:
            return redirect('/edit_client/' + str(client_id))


def delete_client(request, client_id):
    client = Client.objects.get(admin=client_id)
    try:
        client.delete()
        messages.success(request, "Client supprimé avec succès.")
        return redirect('manage_client')
    except:
        messages.error(request, "Échec de la suppression du client.")
        return redirect('manage_client')
    

def add_farmer(request):
    form = AgriculteurForm()
    context = {
        "form": form
    }
    return render(request, 'admin/add_farmer_template.html', context)
def add_farmer_save(request):
    if request.method != "POST":
        messages.error(request, "Méthode invalide")
        return redirect('add_farmer')
    else:
        form = AgriculteurForm(request.POST, request.FILES)

        if form.is_valid():
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']
            username = form.cleaned_data['username']
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            address = form.cleaned_data['address']
            phone = form.cleaned_data['phone']
            gender = form.cleaned_data['gender']

            if len(request.FILES) != 0:
                profile_pic = request.FILES['profile_pic']
                fs = FileSystemStorage()
                filename = fs.save(profile_pic.name, profile_pic)
                profile_pic_url = fs.url(filename)
            else:
                profile_pic_url = None

            try:
                user = CustomUser.objects.create_user(
                    username=username,
                    password=password,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    user_type=2  # supposons que 2 = agriculteur
                )
                user.agriculteur.address = address
                user.agriculteur.phone = phone
                user.agriculteur.gender = gender
                user.agriculteur.profile_pic = profile_pic_url
                user.save()
                messages.success(request, "Agriculteur ajouté avec succès !")
                return redirect('add_farmer')
            except Exception as e:
                messages.error(request, f"Erreur lors de l'ajout : {e}")
                return redirect('add_farmer')
        else:
            messages.error(request, "Formulaire invalide.")
            return redirect('add_farmer')


def manage_farmer(request):
    farmers = Agriculteur.objects.all()
    context = {
        "farmers": farmers
    }
    return render(request, 'admin/manage_farmer_template.html', context)


def edit_farmer(request, farmer_id):
    request.session['farmer_id'] = farmer_id

    farmer = Agriculteur.objects.get(admin=farmer_id)
    form = AgriculteurForm(instance=farmer.admin)

    form.fields['email'].initial = farmer.admin.email
    form.fields['username'].initial = farmer.admin.username
    form.fields['first_name'].initial = farmer.admin.first_name
    form.fields['last_name'].initial = farmer.admin.last_name
    form.fields['address'].initial = farmer.address
    form.fields['phone'].initial = farmer.phone
    form.fields['gender'].initial = farmer.gender

    context = {
        "id": farmer_id,
        "form": form
    }
    return render(request, "admin/edit_farmer_template.html", context)


def edit_farmer_save(request):
    if request.method != "POST":
        return HttpResponse("Méthode invalide !")
    else:
        farmer_id = request.session.get('farmer_id')
        if farmer_id is None:
            return redirect('manage_farmer')

        farmer = Agriculteur.objects.get(admin=farmer_id)
        form = AgriculteurForm(request.POST, request.FILES, instance=farmer.admin)
        if form.is_valid():
            email = form.cleaned_data['email']
            username = form.cleaned_data['username']
            first_name = form.cleaned_data['first_name']
            last_name = form.cleaned_data['last_name']
            address = form.cleaned_data['address']
            phone = form.cleaned_data['phone']
            gender = form.cleaned_data['gender']

            if len(request.FILES) != 0:
                profile_pic = request.FILES['profile_pic']
                fs = FileSystemStorage()
                filename = fs.save(profile_pic.name, profile_pic)
                profile_pic_url = fs.url(filename)
            else:
                profile_pic_url = None

            try:        
                user = CustomUser.objects.get(id=farmer_id)
                user.first_name = first_name
                user.last_name = last_name
                user.email = email
                user.username = username
                user.save()

                farmer_model = Agriculteur.objects.get(admin=farmer_id)
                farmer_model.address = address
                farmer_model.phone = phone
                farmer_model.gender = gender
                if profile_pic_url:
                    farmer_model.profile_pic = profile_pic_url
                farmer_model.save()

                del request.session['farmer_id']
                messages.success(request, "Agriculteur mis à jour avec succès !")
                return redirect('/edit_farmer/' + str(farmer_id))
            except Exception as e:
                messages.error(request, f"Erreur : {e}")
                return redirect('/edit_farmer/' + str(farmer_id))
        else:
            return redirect('/edit_farmer/' + str(farmer_id))


def delete_farmer(request, farmer_id):
    farmer = Agriculteur.objects.get(admin=farmer_id)
    try:
        farmer.delete()
        messages.success(request, "Agriculteur supprimé avec succès.")
        return redirect('manage_farmer')
    except:
        messages.error(request, "Échec de la suppression de l'agriculteur.")
        return redirect('manage_farmer')

@login_required
def confirm_cash_payment(request, commande_id):
    """
    Confirme un paiement en espèces pour une commande.
    Cette vue est accessible via AJAX ou en requête POST normale.
    """
    if request.method == 'POST':
        try:
            # Récupérer la commande
            commande = get_object_or_404(Commande, id=commande_id)
            
            # Vérifier si l'utilisateur a les droits pour confirmer ce paiement
            if not request.user.is_staff and commande.client != request.user:
                messages.error(request, "Vous n'êtes pas autorisé à effectuer cette action.")
                return JsonResponse({'success': False, 'error': 'Non autorisé'}, status=403)
            
            # Créer un enregistrement de paiement
            paiement = Paiement.objects.create(
                commande=commande,
                amount=commande.total,
                payment_method='a_la_livraison',  # Utilisation de la valeur exacte définie dans les choix
                payment_status='Validated',  # Utilisation de la valeur exacte définie dans les choix
                payment_date=timezone.now()
            )
            
            # Mettre à jour le statut de la commande
            commande.statut = 'payee'
            commande.save()
            
            # Réponse JSON pour les requêtes AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Paiement en espèces confirmé avec succès!',
                    'redirect_url': reverse('liste_commandes')
                })
                
            # Redirection pour les requêtes normales
            messages.success(request, 'Paiement en espèces confirmé avec succès!')
            return redirect('liste_commandes')
            
        except Exception as e:
            error_msg = f"Une erreur s'est produite: {str(e)}"
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg}, status=500)
                
            messages.error(request, error_msg)
            return redirect('liste_commandes')
    
    # Si la méthode n'est pas POST
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)
        
    return redirect('liste_commandes')

@login_required
def create_order(request):
    products = Produit.objects.all()
    if request.method == "POST":
        product_ids = request.POST.getlist('products')  # Les produits sélectionnés
        quantities = request.POST.getlist('quantities')  # Quantités
        client_id = request.POST.get('client') or request.POST.get('client_id')

        # Vérifier le client cible (si un gérant crée pour un client)
        if request.user.is_staff or request.user.is_superuser:
            if not client_id:
                messages.error(request, "Veuillez sélectionner un client pour cette commande.")
                return render(request, 'admin/create_order.html', {'produits': products})
            try:
                from .models import CustomUser
                client_user = CustomUser.objects.get(id=client_id, role__iexact='Client')
            except CustomUser.DoesNotExist:
                messages.error(request, "Client introuvable ou invalide.")
                return render(request, 'admin/create_order.html', {'produits': products})
        else:
            # Client connecté (parcours client)
            client_user = request.user

        # Créer la commande rattachée au bon client
        order = Commande(client=client_user)
        order.save()

        total_amount = 0
        for product_id, quantity in zip(product_ids, quantities):
            product = Produit.objects.get(id=product_id)
            qte = int(quantity)
            sous_total = product.prix_unitaire * qte
            total_amount += sous_total
            LigneCommande.objects.create(commande=order, produit=product, quantite=qte, sous_total=sous_total)

        # Enregistrer le total dans le champ 'total'
        order.total = total_amount
        order.save()

        messages.success(request, "Commande créée avec succès !")
        return redirect('liste_commandes')

    return render(request, 'admin/create_order.html', {'produits': products})

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def commandes_en_attente(request):
    """
    Affiche la liste des commandes en attente de traitement.
    Cette vue est accessible uniquement aux administrateurs.
    """
    # Récupérer toutes les commandes en attente
    commandes = Commande.objects.filter(statut='en_attente').order_by('-date_commande')
    
    context = {
        'commandes': commandes,
        'titre': 'Commandes en attente',
    }
    
    return render(request, 'admin/commandes_en_attente.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def liste_stock(request):
    """
    Affiche la liste des produits en stock.
    Cette vue est accessible uniquement aux administrateurs.
    """
    # Récupérer tous les produits
    produits = Produit.objects.all().order_by('nom_produit')
    
    # Identifier les produits en alerte de stock
    for produit in produits:
        produit.en_alerte = produit.est_en_alerte()
    
    context = {
        'produits': produits,
        'titre': 'État du stock',
    }
    
    return render(request, 'admin/liste_stock.html', context)
@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def ajouter_mouvement(request):
    """
    Ajoute un mouvement de stock (entrée ou sortie).
    Cette vue est accessible uniquement aux administrateurs.
    """
    if request.method == 'POST':
        form = MouvementStockForm(request.POST)
        if form.is_valid():
            mouvement = form.save(commit=False)
            
            # Mettre à jour le stock du produit
            produit = mouvement.produit
            if mouvement.type_mouvement == 'ENTREE':
                produit.update_stock(mouvement.quantite)
            else:  # SORTIE
                try:
                    produit.update_stock(-mouvement.quantite)
                except ValueError as e:
                    messages.error(request, str(e))
                    return redirect('ajouter_mouvement')
            
            mouvement.save()
            messages.success(request, "Mouvement de stock enregistré avec succès.")
            return redirect('liste_stock')
    else:
        form = MouvementStockForm()
    
    context = {
        'form': form,
        'titre': 'Ajouter un mouvement de stock',
    }
    
    return render(request, 'admin/ajouter_mouvements.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def exporter_stock_excel(request):
    """
    Exporte l'état du stock au format Excel.
    Cette vue est accessible uniquement aux administrateurs.
    """
    produits = Produit.objects.all().order_by('nom_produit')
    
    # Créer un classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "État du stock"
    
    # Ajouter les en-têtes
    headers = ['Produit', 'Type', 'Quantité', 'Prix unitaire', 'Montant total', 'Statut']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
    
    # Ajouter les données
    for row_num, produit in enumerate(produits, 2):
        ws[f'A{row_num}'] = produit.nom_produit
        ws[f'B{row_num}'] = produit.type_produit
        ws[f'C{row_num}'] = float(produit.quantite)
        ws[f'D{row_num}'] = float(produit.prix_unitaire)
        ws[f'E{row_num}'] = float(produit.montant_total)
        ws[f'F{row_num}'] = "En alerte" if produit.est_en_alerte() else "Normal"
    
    # Créer la réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=etat_stock.xlsx'
    wb.save(response)
    
    return response

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def ajouter_client(request):
    """
    Affiche et traite le formulaire d'ajout d'un nouveau client.
    Cette fonction est accessible uniquement aux administrateurs.
    """
    print("=== DEBUG: Entrée dans la vue ajouter_client ===")  # Debug
    is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    print(f"=== DEBUG: is_ajax = {is_ajax}")  # Debug
    template_name = 'admin/ajouter_client.html'
    form_template = 'admin/_ajouter_client_form.html' if is_ajax else template_name

    if request.method == 'POST':
        print("=== DEBUG: Requête POST reçue")  # Debug
        print(f"=== DEBUG: Données POST: {request.POST}")  # Debug
        print(f"=== DEBUG: Fichiers: {request.FILES}")  # Debug
        form = ClientForm(request.POST, request.FILES)
        print(f"=== DEBUG: Formulaire valide: {form.is_valid()}")  # Debug
        if not form.is_valid():
            print(f"=== DEBUG: Erreurs du formulaire: {form.errors}")  # Debug
        if form.is_valid():
            try:
                # Sauvegarder l'utilisateur et le profil client
                user = form.save()
                
                # Mettre à jour le profil client avec les champs supplémentaires
                client = user.client_profile
                client.city = form.cleaned_data.get('city')
                client.postal_code = form.cleaned_data.get('postal_code')
                client.country = form.cleaned_data.get('country')
                client.phone_number_alt = form.cleaned_data.get('phone_number_alt')
                client.date_of_birth = form.cleaned_data.get('date_of_birth')
                client.notes = form.cleaned_data.get('notes')
                client.save()
                
                # Message de succès
                success_message = f"Le client {user.get_full_name()} a été ajouté avec succès."
                
                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'message': success_message,
                        'client_id': client.id,
                        'client_name': str(client)
                    })
                    
                messages.success(request, success_message)
                return redirect('liste_clients')
                
            except Exception as e:
                error_message = f"Une erreur est survenue lors de l'ajout du client : {str(e)}"
                if is_ajax:
                    return JsonResponse({
                        'success': False,
                        'message': error_message
                    }, status=400)
                messages.error(request, error_message)
        else:
            if is_ajax:
                html = render_to_string(form_template, {'form': form}, request=request)
                return JsonResponse({
                    'success': False,
                    'html': html,
                    'message': 'Veuillez corriger les erreurs ci-dessous.'
                }, status=400)
    else:
        form = ClientForm()

    # Si c'est une requête AJAX GET, renvoyer uniquement le formulaire
    if is_ajax and request.method == 'GET':
        html = render_to_string(form_template, {'form': form}, request=request)
        return JsonResponse({'html': html})

    # Pour une requête normale, afficher la page complète
    context = {
        'form': form,
        'titre': 'Ajouter un nouveau client',
        'active_menu': 'clients',
    }
    return render(request, template_name, context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def toggle_activation_client(request, client_id):
    """
    Active ou désactive un client en fonction de son état actuel.
    Cette fonction est utilisée pour basculer l'état d'activation d'un client.
    """
    client = get_object_or_404(Client, id=client_id)
    user = client.admin  # Correction: utiliser le champ admin au lieu de user
    
    # Inverser l'état d'activation
    user.is_active = not user.is_active
    user.save()
    
    status = "activé" if user.is_active else "désactivé"
    messages.success(request, f"Le client {user.get_full_name()} a été {status}.")
    
    return redirect("liste_clients")

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def ajouter_agriculteur(request):
    """
    Affiche et traite le formulaire d'ajout d'un nouvel agriculteur.
    Cette fonction est accessible uniquement aux administrateurs.
    """
    if request.method == 'POST':
        form = AgriculteurForm(request.POST, request.FILES)
        if form.is_valid():
            # Créer un nouvel utilisateur
            user = CustomUser.objects.create_user(
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password'],
                email=form.cleaned_data['email'],
                first_name=form.cleaned_data['first_name'],
                last_name=form.cleaned_data['last_name'],
                role='Agriculteur'
            )
            
            # Créer le profil agriculteur associé
            agriculteur = Agriculteur.objects.get(user=user)
            agriculteur.adresse = form.cleaned_data.get('adresse', '')
            agriculteur.telephone = form.cleaned_data.get('telephone', '')
            agriculteur.specialite = form.cleaned_data.get('specialite', '')
            
            # Traiter la photo de profil si fournie
            if 'photo_profil' in request.FILES:
                agriculteur.photo_profil = request.FILES['photo_profil']
            
            agriculteur.save()
            
            messages.success(request, f"L'agriculteur {user.get_full_name()} a été ajouté avec succès.")
            return redirect('liste_agriculteurs')
    else:
        form = AgriculteurForm()
    
    context = {
        'form': form,
        'titre': 'Ajouter un nouvel agriculteur',
    }
    
    return render(request, 'admin/ajouter_agriculteur.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def modifier_agriculteur(request, pk):
    """
    Affiche et traite le formulaire de modification d'un agriculteur existant.
    Cette fonction est accessible uniquement aux administrateurs.
    """
    agriculteur = get_object_or_404(Agriculteur, pk=pk)
    user = agriculteur.user
    
    if request.method == 'POST':
        form = AgriculteurForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            # Mettre à jour l'utilisateur
            user.username = form.cleaned_data['username']
            user.email = form.cleaned_data['email']
            user.first_name = form.cleaned_data['first_name']
            user.last_name = form.cleaned_data['last_name']
            
            # Si un nouveau mot de passe est fourni, le mettre à jour
            if form.cleaned_data.get('password'):
                user.set_password(form.cleaned_data['password'])
            
            user.save()
            
            # Mettre à jour le profil agriculteur
            agriculteur.adresse = form.cleaned_data.get('adresse', '')
            agriculteur.telephone = form.cleaned_data.get('telephone', '')
            agriculteur.specialite = form.cleaned_data.get('specialite', '')
            
            # Traiter la photo de profil si fournie
            if 'photo_profil' in request.FILES:
                agriculteur.photo_profil = request.FILES['photo_profil']
            
            agriculteur.save()
            
            messages.success(request, f"L'agriculteur {user.get_full_name()} a été modifié avec succès.")
            return redirect('liste_agriculteurs')
    else:
        # Pré-remplir le formulaire avec les données existantes
        initial_data = {
            'username': user.username,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'adresse': agriculteur.adresse,
            'telephone': agriculteur.telephone,
            'specialite': agriculteur.specialite,
        }
        form = AgriculteurForm(initial=initial_data)
    
    context = {
        'form': form,
        'agriculteur': agriculteur,
        'titre': f"Modifier l'agriculteur {user.get_full_name()}",
    }
    
    return render(request, 'admin/modifier_agriculteur.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def supprimer_agriculteur(request, pk):
    """
    Supprime un agriculteur du système.
    Cette fonction est accessible uniquement aux administrateurs.
    """
    agriculteur = get_object_or_404(Agriculteur, pk=pk)
    user = agriculteur.user
    nom_complet = user.get_full_name()
    
    try:
        # La suppression de l'utilisateur entraînera la suppression de l'agriculteur associé
        # grâce à la relation CASCADE
        user.delete()
        messages.success(request, f"L'agriculteur {nom_complet} a été supprimé avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression de l'agriculteur : {e}")
    
    return redirect('liste_agriculteurs')

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def modifier_commande(request, commande_id):
    """
    Affiche et traite le formulaire de modification d'une commande existante.
    Cette vue est accessible uniquement aux administrateurs.
    """
    try:
        commande = get_object_or_404(Commande, id=commande_id)
        
        if request.method == 'POST':
            # Récupérer les données du formulaire
            statut = request.POST.get('statut')
            notes = request.POST.get('notes', '')
            
            # Mettre à jour la commande
            commande.statut = statut
            commande.notes = notes
            commande.save()
            
            messages.success(request, 'La commande a été mise à jour avec succès.')
            return redirect('liste_commandes')
            
        # Préparer le contexte pour le formulaire
        context = {
            'commande': commande,
            'title': f'Modifier la commande #{commande.id}',
            'statuts_commande': Commande.STATUT_CHOICES,
        }
        
        return render(request, 'admin/modifier_commande.html', context)
        
    except Exception as e:
        messages.error(request, f'Une erreur est survenue : {str(e)}')
        return redirect('liste_commandes')

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def liste_commandes_pro(request):
    """
    Vue pour le tableau de commandes professionnel moderne.
    - Interface ultra-professionnelle 2025
    - Compatible DataTables
    - Actions intelligentes selon le statut
    """
    from django.db import models
    
    # Récupérer toutes les commandes avec leurs détails
    commandes = Commande.objects.select_related('client').prefetch_related('details__produit').all().order_by('-date_commande')
    
    # Calculer les statistiques
    stats = {
        'total_commandes': commandes.count(),
        'commandes_validees': commandes.filter(statut='validee').count(),
        'commandes_en_attente': commandes.filter(statut='en_attente').count(),
        'commandes_en_preparation': commandes.filter(statut='en_preparation').count(),
        'commandes_annulees': commandes.filter(statut='annulee').count(),
        'total_ventes': commandes.filter(statut__in=['validee', 'en_preparation', 'livree', 'payee']).aggregate(
            total=models.Sum('total')
        )['total'] or 0,
    }
    
    # Filtrer par client si spécifié
    client_id = request.GET.get('client')
    if client_id:
        commandes = commandes.filter(client_id=client_id)
    
    # Filtrer par statut si spécifié
    statut = request.GET.get('statut')
    if statut:
        commandes = commandes.filter(statut=statut)
    
    # Filtrer par date si spécifié
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    if date_debut:
        commandes = commandes.filter(date_commande__date__gte=date_debut)
    if date_fin:
        commandes = commandes.filter(date_commande__date__lte=date_fin)
    
    context = {
        'commandes': commandes,
        'stats': stats,
        'clients': CustomUser.objects.filter(role='Client'),
        'statuts_choices': Commande.STATUTS,
    }
    
    return render(request, 'admin/liste_commandes_pro.html', context)

def liste_commandes(request):
    """
    Affiche la liste de toutes les commandes avec les détails des produits.
    Cette vue est accessible uniquement aux administrateurs.
    """
    # Récupérer toutes les commandes
    commandes = Commande.objects.all().order_by('-date_commande')
    
    # Filtres
    statut = request.GET.get('statut')
    if statut:
        commandes = commandes.filter(statut=statut)
    client_id = request.GET.get('client')
    if client_id:
        commandes = commandes.filter(client_id=client_id)
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    if date_debut:
        try:
            commandes = commandes.filter(date_commande__date__gte=date_debut)
        except Exception:
            pass
    if date_fin:
        try:
            commandes = commandes.filter(date_commande__date__lte=date_fin)
        except Exception:
            pass
    q = request.GET.get('q', '').strip()
    if q:
        commandes = commandes.filter(
            Q(client__first_name__icontains=q) |
            Q(client__last_name__icontains=q) |
            Q(client__username__icontains=q) |
            Q(client__email__icontains=q) |
            Q(id__icontains=q)
        )
    
    # Récupérer les détails des commandes avec les produits
    commandes = commandes.prefetch_related('details__produit')
    
    # Pagination
    paginator = Paginator(commandes, 10)  # 10 commandes par page
    page = request.GET.get('page')
    commandes_page = paginator.get_page(page)
    
    # Liste des clients pour le filtre
    clients = CustomUser.objects.filter(role__iexact='Client')
    
    # Calculer les statistiques sur le jeu filtré
    total_commandes = commandes.count()
    commandes_en_attente = commandes.filter(statut='en_attente').count()
    commandes_terminees = commandes.filter(statut='terminee').count()
    revenus_totaux = commandes.filter(statut='terminee').aggregate(Sum('total'))['total__sum'] or 0
    
    # Calculer la moyenne par commande terminée
    moyenne_par_commande = revenus_totaux / commandes_terminees if commandes_terminees > 0 else 0
    
    # Compter les notifications non lues pour l'admin
    from agricole.models import Notification
    notifications_count = Notification.objects.filter(
        user=request.user,
        is_read=False
    ).count()
    
    # Préparer les détails des produits pour chaque commande
    for commande in commandes_page:
        commande.produits_details = [
            f"{detail.quantite} x {detail.produit.nom_produit if detail.produit else 'Produit supprimé'}"
            for detail in commande.details.all()
        ]
    
    context = {
        'commandes': commandes_page,
        'clients': clients,
        'titre': 'Liste des commandes',
        'statut_filtre': statut or '',
        'client_filtre': client_id or '',
        'date_debut': date_debut or '',
        'date_fin': date_fin or '',
        'q': q,
        'total_commandes': total_commandes,
        'commandes_en_attente': commandes_en_attente,
        'commandes_terminees': commandes_terminees,
        'revenus_totaux': revenus_totaux,
        'moyenne_par_commande': moyenne_par_commande,
        'notifications_count': notifications_count,
    }
    
    return render(request, 'admin/liste_commandes.html', context)


@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def detail_commande_modal(request, commande_id):
    """
    Vue pour afficher les détails d'une commande dans un modal (AJAX).
    Retourne un fragment HTML à afficher dans le modal.
    """
    commande = get_object_or_404(
        Commande.objects.select_related('client').prefetch_related('details__produit'),
        id=commande_id
    )
    
    return render(request, 'admin/partials/commande_detail_modal_content.html', {
        'commande': commande
    })


@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def renvoyer_email_paiement(request, commande_id):
    """
    Vue pour renvoyer l'email de paiement au client.
    Accessible uniquement aux admins/superusers.
    """
    commande = get_object_or_404(Commande, id=commande_id)
    
    if commande.statut != 'en_attente_paiement':
        messages.error(request, f"La commande #{commande.id} n'est pas en attente de paiement.")
        return redirect('liste_commandes')
    
    try:
        from django.core.mail import send_mail
        from django.conf import settings
        
        client = commande.client
        
        sujet = f"Rappel : Paiement requis pour votre commande #{commande.id} - Kayupe Agriculture"
        message_html = f"""
        <html>
        <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%); color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0;">
                <h1 style="margin: 0; font-size: 28px;">Kayupe Agriculture</h1>
                <p style="margin: 10px 0 0 0; opacity: 0.9;">Rappel de paiement</p>
            </div>
            
            <div style="background: white; padding: 40px; border: 1px solid #e5e7eb; border-radius: 0 0 10px 10px;">
                <h2 style="color: #1f2937; margin-bottom: 20px;">Rappel : Paiement requis</h2>
                
                <p style="color: #6b7280; font-size: 16px; line-height: 1.6;">
                    Bonjour {client.get_full_name() or client.username},
                </p>
                
                <p style="color: #6b7280; font-size: 16px; line-height: 1.6;">
                    Ceci est un rappel concernant votre commande <strong>#{commande.id}</strong> qui est toujours en attente de paiement.
                </p>
                
                <div style="background: #fef2f2; border-left: 4px solid #ef4444; padding: 20px; margin: 25px 0;">
                    <h3 style="color: #991b1b; margin-top: 0; margin-bottom: 15px;">Détails de la commande :</h3>
                    <ul style="color: #991b1b; line-height: 1.8; padding-left: 20px;">
                        <li><strong>Numéro de commande :</strong> #{commande.id}</li>
                        <li><strong>Date :</strong> {commande.date_commande.strftime('%d/%m/%Y à %H:%M')}</li>
                        <li><strong>Montant à payer :</strong> {commande.total:,} FCFA</li>
                        <li><strong>Statut :</strong> En attente de paiement</li>
                    </ul>
                </div>
                
                <div style="text-align: center; margin: 30px 0;">
                    <a href="{settings.SITE_URL if hasattr(settings, 'SITE_URL') else 'http://127.0.0.1:8000'}/paiement/commande/{commande.id}/" 
                       style="background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%); color: white; padding: 15px 30px; text-decoration: none; border-radius: 8px; font-weight: bold; display: inline-block;">
                        Payer ma commande maintenant
                    </a>
                </div>
                
                <p style="color: #6b7280; font-size: 16px; line-height: 1.6;">
                    Une fois le paiement effectué, votre commande sera automatiquement validée et mise en préparation.
                </p>
                
                <div style="border-top: 1px solid #e5e7eb; padding-top: 20px; margin-top: 30px; text-align: center;">
                    <p style="color: #9ca3af; font-size: 14px; margin: 0;">
                        Merci de votre confiance !<br>
                        L'équipe Kayupe Agriculture
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        send_mail(
            sujet,
            f"Rappel : Votre commande #{commande.id} est en attente de paiement. Veuillez la régler pour finalisation.",
            settings.DEFAULT_FROM_EMAIL,
            [client.email],
            html_message=message_html,
            fail_silently=False,
        )
        
        messages.success(
            request,
            f"L'email de rappel a été envoyé avec succès à {client.email}."
        )
        
    except Exception as e:
        messages.error(
            request,
            f"Erreur lors de l'envoi de l'email : {str(e)}"
        )
    
    return redirect('liste_commandes')


@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def ajouter_commande(request):
    """
    Vue pour ajouter une nouvelle commande côté administrateur.
    - Accessible uniquement aux admins/superusers
    - La commande est validée automatiquement
    - Le stock est mis à jour
    - Une notification est envoyée au client
    """
    if request.method == 'POST':
        # Récupérer les données du formulaire
        client_id = request.POST.get('client')
        date_commande = request.POST.get('date_commande')
        produits_ids = request.POST.getlist('produits[]')
        quantites = request.POST.getlist('quantite[]')

        # Debug: afficher les données reçues
        print(f"Données reçues - client_id: {client_id}, produits: {produits_ids}, quantites: {quantites}")
        
        # Vérifier si le client_id existe
        if client_id:
            try:
                client_test = CustomUser.objects.get(id=client_id)
                print(f"Client trouvé: {client_test.username} (Role: {client_test.role})")
            except CustomUser.DoesNotExist:
                print(f"Aucun client trouvé avec l'ID: {client_id}")
                # Afficher tous les utilisateurs pour débogage
                all_users = CustomUser.objects.all()
                for user in all_users:
                    print(f"Utilisateur: {user.username} (ID: {user.id}, Role: {user.role})")

        # Validation des données
        if not client_id:
            messages.error(request, "Veuillez sélectionner un client.")
            return redirect('ajouter_commande')
        
        if not produits_ids or not quantites or len(produits_ids) != len(quantites):
            messages.error(request, "Veuillez ajouter au moins un produit à la commande.")
            return redirect('ajouter_commande')

        try:
            with transaction.atomic():
                # Récupérer le client
                # Utiliser les constantes du modèle CustomUser
                CLIENT_ROLE = CustomUser.CLIENT  # '3'
                
                # Essayer différentes méthodes pour trouver le client
                client = None
                error_messages = []
                
                try:
                    client = CustomUser.objects.get(id=client_id, role=CLIENT_ROLE)
                    print(f"Client trouvé via role=CLIENT_ROLE ('{CLIENT_ROLE}')")
                except CustomUser.DoesNotExist:
                    error_messages.append(f"role=CLIENT_ROLE ('{CLIENT_ROLE}')")
                
                if not client:
                    try:
                        client = CustomUser.objects.get(id=client_id, user_type=CLIENT_ROLE)
                        print(f"Client trouvé via user_type=CLIENT_ROLE ('{CLIENT_ROLE}')")
                    except CustomUser.DoesNotExist:
                        error_messages.append(f"user_type=CLIENT_ROLE ('{CLIENT_ROLE}')")
                
                if not client:
                    try:
                        client = CustomUser.objects.get(id=client_id, role='Client')
                        print("Client trouvé via role='Client'")
                    except CustomUser.DoesNotExist:
                        error_messages.append("role='Client'")
                
                if not client:
                    # En dernier recours, accepter n'importe quel utilisateur avec l'ID spécifié
                    try:
                        client = CustomUser.objects.get(id=client_id)
                        print(f"Client trouvé via ID seul (débogage): {client.username} (Role: {client.role})")
                        messages.warning(request, f"Utilisateur '{client.username}' trouvé mais sans rôle de client valide. Commande créée pour débogage.")
                    except CustomUser.DoesNotExist:
                        # Afficher tous les utilisateurs pour débogage
                        all_users = CustomUser.objects.all()
                        print(f"=== DÉBOGAGE POST: Recherche client ID {client_id} ===")
                        for user in all_users:
                            print(f"ID: {user.id}, Username: {user.username}, Role: '{user.role}', User Type: '{user.user_type}'")
                        
                        raise CustomUser.DoesNotExist(f"Aucun utilisateur trouvé avec ID {client_id}. Tentatives avec: {', '.join(error_messages)}")
                
                # Création de la commande
                commande = Commande.objects.create(
                    client=client,
                    date_commande=timezone.now() if not date_commande else timezone.now(),
                    statut='en_attente_paiement'  # en attente de paiement client
                )

                # Créer les détails de commande depuis les tableaux postés
                total = 0
                for pid, qte in zip(produits_ids, quantites):
                    if not pid or not qte:
                        continue
                    
                    produit = Produit.objects.get(id=pid)
                    qte_int = int(qte)
                    
                    if qte_int <= 0:
                        raise Exception("La quantité doit être supérieure à 0.")
                    
                    # Vérifier stock
                    if produit.quantite < qte_int:
                        raise Exception(f"Stock insuffisant pour le produit {produit.nom_produit}. Stock disponible: {produit.quantite}")

                    # Prix unitaire courant
                    prix_unitaire = produit.prix_unitaire
                    DetailCommande.objects.create(
                        commande=commande,
                        produit=produit,
                        quantite=qte_int,
                        prix_unitaire=prix_unitaire,
                    )

                    # Mises à jour total et stock
                    total += qte_int * prix_unitaire
                    produit.quantite -= qte_int
                    produit.save()
                    
                    # Créer un mouvement de stock
                    MouvementStock.objects.create(
                        produit=produit,
                        type_mouvement=MouvementStock.SORTIE,
                        quantite=qte_int,
                        commentaire=f"Commande #{commande.id} - {client.get_full_name()}"
                    )
                
                commande.total = total
                commande.save()

                # Créer une notification pour le client
                notification = Notification.objects.create(
                    user=client,
                    message=f"Votre commande #{commande.id} est en attente de paiement. Veuillez la régler pour finalisation.",
                    type='warning',
                    url=f"/paiement/commande/{commande.id}/"
                )

                # Envoyer un email au client pour paiement
                try:
                    from django.core.mail import send_mail
                    from django.conf import settings
                    
                    sujet = f"Action requise : Paiement de votre commande #{commande.id} - Kayupe Agriculture"
                    message_html = f"""
                    <html>
                    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                        <div style="background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%); color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0;">
                            <h1 style="margin: 0; font-size: 28px;">Kayupe Agriculture</h1>
                            <p style="margin: 10px 0 0 0; opacity: 0.9;">Votre partenaire agricole de confiance</p>
                        </div>
                        
                        <div style="background: white; padding: 40px; border: 1px solid #e5e7eb; border-radius: 0 0 10px 10px;">
                            <h2 style="color: #1f2937; margin-bottom: 20px;">Action requise : Paiement de votre commande</h2>
                            
                            <p style="color: #6b7280; font-size: 16px; line-height: 1.6;">
                                Bonjour {client.get_full_name() or client.username},
                            </p>
                            
                            <p style="color: #6b7280; font-size: 16px; line-height: 1.6;">
                                Le gérant a créé votre commande <strong>#{commande.id}</strong>. 
                                Pour finaliser votre commande, veuillez procéder au paiement.
                            </p>
                            
                            <div style="background: #fef3c7; border-left: 4px solid #f59e0b; padding: 20px; margin: 25px 0;">
                                <h3 style="color: #92400e; margin-top: 0; margin-bottom: 15px;">Détails de la commande :</h3>
                                <ul style="color: #92400e; line-height: 1.8; padding-left: 20px;">
                                    <li><strong>Numéro de commande :</strong> #{commande.id}</li>
                                    <li><strong>Date :</strong> {commande.date_commande.strftime('%d/%m/%Y à %H:%M')}</li>
                                    <li><strong>Montant à payer :</strong> {commande.total:,} FCFA</li>
                                    <li><strong>Statut :</strong> En attente de paiement</li>
                                </ul>
                            </div>
                            
                            <div style="background: #ecfdf5; border-left: 4px solid #10b981; padding: 20px; margin: 25px 0;">
                                <p style="color: #065f46; margin: 0;">
                                    <strong>Produits commandés :</strong>
                                </p>
                                <ul style="color: #065f46; margin: 10px 0 0 20px;">
                    """
                    
                    # Ajouter les produits de la commande
                    for detail in commande.details.all():
                        message_html += f"<li>{detail.quantite} x {detail.produit.nom_produit} - {detail.prix_unitaire:,} FCFA</li>"
                    
                    message_html += f"""
                                </ul>
                            </div>
                            
                            <div style="text-align: center; margin: 30px 0;">
                                <a href="{settings.SITE_URL if hasattr(settings, 'SITE_URL') else 'http://127.0.0.1:8000'}/paiement/commande/{commande.id}/" 
                                   style="background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%); color: white; padding: 15px 30px; text-decoration: none; border-radius: 8px; font-weight: bold; display: inline-block;">
                                    Payer ma commande maintenant
                                </a>
                            </div>
                            
                            <p style="color: #6b7280; font-size: 16px; line-height: 1.6;">
                                Une fois le paiement effectué, votre commande sera automatiquement validée et mise en préparation.
                            </p>
                            
                            <div style="border-top: 1px solid #e5e7eb; padding-top: 20px; margin-top: 30px; text-align: center;">
                                <p style="color: #9ca3af; font-size: 14px; margin: 0;">
                                    Merci de votre confiance !<br>
                                    L'équipe Kayupe Agriculture
                                </p>
                            </div>
                        </div>
                    </body>
                    </html>
                    """
                    
                    send_mail(
                        sujet,
                        f"Votre commande #{commande.id} est en attente de paiement. Veuillez la régler pour finalisation.",
                        settings.DEFAULT_FROM_EMAIL,
                        [client.email],
                        html_message=message_html,
                        fail_silently=False,
                    )
                    
                    # Notifier l'admin que l'email a été envoyé
                    messages.info(
                        request,
                        f"Un email de demande de paiement a été envoyé à {client.email}."
                    )
                    
                except Exception as e:
                    # Si l'email échoue, on ne bloque pas le processus
                    messages.warning(
                        request,
                        f"La commande a été créée mais l'email de demande de paiement n'a pas pu être envoyé: {str(e)}"
                    )

                messages.success(
                    request,
                    f"La commande #{commande.id} a été créée avec succès. "
                    f"Le client {client.get_full_name()} a reçu un email pour procéder au paiement. "
                    f"La commande sera validée automatiquement après paiement."
                )

            return redirect('liste_commandes')

        except CustomUser.DoesNotExist:
            messages.error(request, "Le client sélectionné n'existe pas ou n'a pas le rôle de client.")
        except Produit.DoesNotExist:
            messages.error(request, "Un des produits sélectionnés n'existe pas.")
        except Exception as e:
            messages.error(
                request,
                f"Une erreur est survenue lors de la création de la commande : {str(e)}"
            )
        
        return redirect('ajouter_commande')
    
    # GET - Afficher le formulaire
    form = CommandeForm()

    # Liste des clients pour autocomplétion
    # Afficher TOUS les utilisateurs pour débogage complet
    all_users = CustomUser.objects.all()
    print(f"=== DÉBOGAGE COMPLET: Tous les utilisateurs ({all_users.count()})) ===")
    for user in all_users:
        print(f"ID: {user.id}, Username: {user.username}, Role: '{user.role}', User Type: '{user.user_type}', Email: {user.email}, is_active: {user.is_active}")
    
    # Utiliser les constantes du modèle CustomUser
    CLIENT_ROLE = CustomUser.CLIENT  # '3'
    
    # Essayer différentes méthodes pour trouver les clients
    clients_role = CustomUser.objects.filter(role=CLIENT_ROLE)
    clients_user_type = CustomUser.objects.filter(user_type=CLIENT_ROLE)
    clients_role_text = CustomUser.objects.filter(role='Client')
    clients_role_num = CustomUser.objects.filter(role='3')
    clients_user_type_num = CustomUser.objects.filter(user_type='3')
    
    print(f"\n=== ANALYSE DES FILTRES ===")
    print(f"Clients avec role=CLIENT_ROLE ('{CLIENT_ROLE}'): {clients_role.count()}")
    print(f"Clients avec user_type=CLIENT_ROLE ('{CLIENT_ROLE}'): {clients_user_type.count()}")
    print(f"Clients avec role='Client': {clients_role_text.count()}")
    print(f"Clients avec role='3': {clients_role_num.count()}")
    print(f"Clients avec user_type='3': {clients_user_type_num.count()}")
    
    # Afficher les détails de chaque filtre
    print(f"\n=== DÉTAILS DES FILTRES ===")
    print("Filtre role=CLIENT_ROLE:")
    for client in clients_role:
        print(f"  - {client.username} (ID: {client.id}, Role: '{client.role}', User Type: '{client.user_type}')")
    
    print("Filtre user_type=CLIENT_ROLE:")
    for client in clients_user_type:
        print(f"  - {client.username} (ID: {client.id}, Role: '{client.role}', User Type: '{client.user_type}')")
    
    print("Filtre role='3':")
    for client in clients_role_num:
        print(f"  - {client.username} (ID: {client.id}, Role: '{client.role}', User Type: '{client.user_type}')")
    
    print("Filtre user_type='3':")
    for client in clients_user_type_num:
        print(f"  - {client.username} (ID: {client.id}, Role: '{client.role}', User Type: '{client.user_type}')")
    
    # Utiliser la méthode qui trouve le plus de clients
    max_count = 0
    best_clients = CustomUser.objects.none()
    best_method = "aucun"
    
    if clients_role.count() > max_count:
        max_count = clients_role.count()
        best_clients = clients_role
        best_method = f"role=CLIENT_ROLE ('{CLIENT_ROLE}')"
    
    if clients_user_type.count() > max_count:
        max_count = clients_user_type.count()
        best_clients = clients_user_type
        best_method = f"user_type=CLIENT_ROLE ('{CLIENT_ROLE}')"
    
    if clients_role_text.count() > max_count:
        max_count = clients_role_text.count()
        best_clients = clients_role_text
        best_method = "role='Client'"
    
    if clients_role_num.count() > max_count:
        max_count = clients_role_num.count()
        best_clients = clients_role_num
        best_method = "role='3'"
    
    if clients_user_type_num.count() > max_count:
        max_count = clients_user_type_num.count()
        best_clients = clients_user_type_num
        best_method = "user_type='3'"
    
    # Si le meilleur filtre trouve moins de clients que attendu, utiliser tous les utilisateurs actifs
    if max_count < 10:  # Si on trouve moins de 10 clients (vous avez dit qu'il y en a 13)
        all_users = CustomUser.objects.all()  # TOUS les utilisateurs, pas seulement les actifs
        all_active_users = CustomUser.objects.filter(is_active=True)
        print(f"\n⚠️  PEU DE CLIENTS TROUVÉS ({max_count}) alors qu'il y en a 13 au total")
        print(f"TOUS les utilisateurs: {all_users.count()}")
        print(f"Utilisateurs actifs: {all_active_users.count()}")
        
        # Analyser les différences
        print(f"\n=== ANALYSE DES DIFFÉRENCES ===")
        print(f"Clients trouvés par le meilleur filtre ({best_method}):")
        found_ids = set()
        for client in best_clients:
            print(f"  ✓ {client.username} (ID: {client.id}, Role: '{client.role}', User Type: '{client.user_type}', Actif: {client.is_active})")
            found_ids.add(client.id)
        
        print(f"\nClients manquants (non trouvés par le filtre):")
        missing_count = 0
        for user in all_users:
            if user.id not in found_ids:
                print(f"  ✗ {user.username} (ID: {user.id}, Role: '{user.role}', User Type: '{user.user_type}', Actif: {user.is_active})")
                missing_count += 1
        
        print(f"\nRésumé: {max_count} clients trouvés, {missing_count} manquants, {all_users.count()} au total")
        
        # Afficher TOUS les utilisateurs pour être sûr de tout voir
        clients = all_users
        best_method = "TOUS les utilisateurs (débogage complet)"
        messages.warning(request, f"Seulement {max_count} clients trouvés avec les filtres standards sur {all_users.count()} au total. Affichage de TOUS les utilisateurs pour débogage.")
    else:
        clients = best_clients
    
    print(f"\n=== MÉTHODE FINALE CHOISIE: {best_method} ({clients.count()} clients) ===")
    
    # Debug: afficher les clients trouvés
    print(f"Nombre final de clients: {clients.count()}")
    for client in clients:
        print(f"Client FINAL: {client.username} (ID: {client.id}, Role: {client.role}, User Type: {client.user_type})")

    # Liste des produits disponibles (stock normal uniquement)
    # Afficher seulement les produits avec un stock disponible
    produits = (
        Produit.objects
        .filter(quantite__gt=0)
        .order_by('nom_produit')
    )

    context = {
        'form': form,
        'clients': clients,
        'produits': produits,
        'titre': 'Ajouter une nouvelle commande',
    }

    return render(request, 'admin/ajouter_commande.html', context)
@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def export_commandes_pdf(request):
    """
    Exporte la liste des commandes au format PDF.
    Cette vue est accessible uniquement aux administrateurs.
    """
    # Récupérer les commandes avec les mêmes filtres que dans liste_commandes
    commandes = Commande.objects.all().order_by('-date_commande')
    
    statut = request.GET.get('statut')
    if statut:
        commandes = commandes.filter(statut=statut)
    client_id = request.GET.get('client')
    if client_id:
        commandes = commandes.filter(client_id=client_id)
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    if date_debut:
        try:
            commandes = commandes.filter(date_commande__date__gte=date_debut)
        except Exception:
            pass
    if date_fin:
        try:
            commandes = commandes.filter(date_commande__date__lte=date_fin)
        except Exception:
            pass
    q = request.GET.get('q', '').strip()
    if q:
        commandes = commandes.filter(
            Q(client__first_name__icontains=q) |
            Q(client__last_name__icontains=q) |
            Q(client__username__icontains=q) |
            Q(client__email__icontains=q) |
            Q(id__icontains=q)
        )
    
    # Préparer le contexte pour le template
    context = {
        'commandes': commandes,
        'titre': 'Liste des commandes',
        'date': timezone.now().strftime('%d/%m/%Y %H:%M'),
        'statut_filtre': statut,
        'date_debut': date_debut or '',
        'date_fin': date_fin or '',
        'q': q,
    }
    
    # Rendre le template HTML
    template = get_template('admin/commandes_pdf.html')
    html = template.render(context)
    
    # Créer la réponse HTTP avec le PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="liste_commandes.pdf"'
    
    # Générer le PDF à partir du HTML
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    # Retourner la réponse HTTP
    if pisa_status.err:
        return HttpResponse('Une erreur est survenue lors de la génération du PDF.', status=500)
    return response

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def export_commandes_excel(request):
    """
    Exporte la liste des commandes au format Excel.
    Cette vue est accessible uniquement aux administrateurs.
    """
    # Récupérer les commandes avec les mêmes filtres que dans liste_commandes
    commandes = Commande.objects.all().order_by('-date_commande')
    
    statut = request.GET.get('statut')
    if statut:
        commandes = commandes.filter(statut=statut)
    
    client_id = request.GET.get('client')
    if client_id:
        commandes = commandes.filter(client_id=client_id)
    
    # Créer un classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liste des commandes"
    
    # Ajouter les en-têtes
    headers = ['ID', 'Client', 'Date', 'Statut', 'Total', 'Produits']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
    
    # Ajouter les données
    for row_num, commande in enumerate(commandes, 2):
        ws[f'A{row_num}'] = commande.id
        ws[f'B{row_num}'] = str(commande.client)
        ws[f'C{row_num}'] = commande.date_commande.strftime('%d/%m/%Y %H:%M')
        ws[f'D{row_num}'] = commande.get_statut_display()
        ws[f'E{row_num}'] = float(commande.total)
        
        # Liste des produits de la commande
        produits = []
        for detail in commande.detailcommande_set.all():
            produits.append(f"{detail.produit.nom_produit} (x{detail.quantite})")
        ws[f'F{row_num}'] = ", ".join(produits)
    
    # Créer la réponse HTTP avec le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=liste_commandes.xlsx'
    wb.save(response)
    
    return response

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def exporter_stock_pdf(request):
    """
    Exporte l'état du stock au format PDF.
    Cette vue est accessible uniquement aux administrateurs.
    """
    produits = Produit.objects.all().order_by('nom_produit')
    
    # Calculer les totaux
    total_produits = produits.count()
    total_valeur = sum(p.prix_unitaire * p.quantite for p in produits)
    
    # Contexte pour le template
    context = {
        'produits': produits,
        'total_produits': total_produits,
        'total_valeur': total_valeur,
        'date_export': timezone.now().strftime('%d/%m/%Y à %H:%M')
    }
    
    # Rendre le template HTML
    template = get_template('admin/stock_pdf.html')
    html = template.render(context)
    
    # Créer la réponse HTTP avec le PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="etat_stock.pdf"'
    
    # Générer le PDF à partir du HTML
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    # Retourner la réponse HTTP
    if pisa_status.err:
        return HttpResponse('Une erreur est survenue lors de la génération du PDF.', status=500)
    return response

# Equipment Rental Management Views
@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def ajouter_materiel(request):
    """Ajouter un nouveau matériel agricole"""
    if request.method == 'POST':
        nom = request.POST.get('nom')
        categorie = request.POST.get('categorie')
        description = request.POST.get('description')
        prix_location_jour = request.POST.get('prix_location_jour')
        marque = request.POST.get('marque')
        modele = request.POST.get('modele')
        quantite_totale = request.POST.get('quantite_totale', 1)
        
        try:
            materiel = Materiel.objects.create(
                nom=nom,
                categorie=categorie,
                description=description,
                prix_location_jour=prix_location_jour,
                marque=marque,
                modele=modele,
                quantite_totale=quantite_totale,
                quantite_disponible=quantite_totale
            )
            
            messages.success(request, f"Le matériel '{materiel.nom}' a été ajouté avec succès.")
            return redirect('gerant_materiels_agricoles')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de l'ajout du matériel: {str(e)}")
    
    return render(request, 'admin/ajouter_materiel.html')

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def modifier_materiel(request, materiel_id):
    """Modifier un matériel agricole existant"""
    materiel = get_object_or_404(Materiel, id=materiel_id)
    
    if request.method == 'POST':
        nom = request.POST.get('nom')
        categorie = request.POST.get('categorie')
        description = request.POST.get('description')
        prix_location_jour = request.POST.get('prix_location_jour')
        marque = request.POST.get('marque')
        modele = request.POST.get('modele')
        statut = request.POST.get('statut')
        quantite_totale = request.POST.get('quantite_totale')
        
        try:
            materiel.nom = nom
            materiel.categorie = categorie
            materiel.description = description
            materiel.prix_location_jour = prix_location_jour
            materiel.marque = marque
            materiel.modele = modele
            materiel.statut = statut
            materiel.quantite_totale = quantite_totale
            
            # Mettre à jour la disponibilité si nécessaire
            if statut == 'disponible':
                materiel.disponible = True
                materiel.quantite_disponible = quantite_totale
            else:
                materiel.disponible = False
                materiel.quantite_disponible = 0
            
            materiel.save()
            
            messages.success(request, f"Le matériel '{materiel.nom}' a été modifié avec succès.")
            return redirect('gerant_materiels_agricoles')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la modification du matériel: {str(e)}")
    
    context = {
        'materiel': materiel,
        'categories': Materiel.CATEGORIES,
        'statuts': Materiel.STATUTS
    }
    return render(request, 'admin/modifier_materiel.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def gerer_disponibilite_materiel(request, materiel_id):
    """Gérer la disponibilité d'un matériel"""
    materiel = get_object_or_404(Materiel, id=materiel_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'rendre_disponible':
            materiel.statut = 'disponible'
            materiel.disponible = True
            materiel.quantite_disponible = materiel.quantite_totale
            messages.success(request, f"Le matériel '{materiel.nom}' est maintenant disponible.")
            
        elif action == 'rendre_indisponible':
            materiel.statut = 'indisponible'
            materiel.disponible = False
            materiel.quantite_disponible = 0
            messages.success(request, f"Le matériel '{materiel.nom}' est maintenant indisponible.")
            
        elif action == 'mettre_en_maintenance':
            materiel.statut = 'maintenance'
            materiel.disponible = False
            materiel.quantite_disponible = 0
            messages.success(request, f"Le matériel '{materiel.nom}' est maintenant en maintenance.")
        
        materiel.save()
        return redirect('gerant_materiels_agricoles')
    
    context = {'materiel': materiel}
    return render(request, 'admin/gerer_disponibilite_materiel.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def liste_reservations_materiel(request):
    """Lister toutes les réservations de matériel"""
    reservations = ReservationMateriel.objects.all().order_by('-date_reservation')
    
    # Filtres
    statut_filtre = request.GET.get('statut')
    materiel_filtre = request.GET.get('materiel')
    
    if statut_filtre:
        reservations = reservations.filter(statut=statut_filtre)
    if materiel_filtre:
        reservations = reservations.filter(materiel_id=materiel_filtre)
    
    # Statistiques
    total_reservations = reservations.count()
    reservations_en_attente = reservations.filter(statut='en_attente').count()
    reservations_validees = reservations.filter(statut='validee').count()
    reservations_en_cours = reservations.filter(statut='en_cours').count()
    
    context = {
        'reservations': reservations,
        'total_reservations': total_reservations,
        'reservations_en_attente': reservations_en_attente,
        'reservations_validees': reservations_validees,
        'reservations_en_cours': reservations_en_cours,
        'statuts': ReservationMateriel.STATUTS,
        'materiels': Materiel.objects.all()
    }
    return render(request, 'admin/liste_reservations_materiel.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def valider_reservation_materiel(request, reservation_id):
    """Valider une réservation de matériel"""
    reservation = get_object_or_404(ReservationMateriel, id=reservation_id)
    
    if request.method == 'POST':
        try:
            # Vérifier la disponibilité
            if not reservation.materiel.is_available_for_dates(reservation.date_debut, reservation.date_fin):
                messages.error(request, f"Le matériel '{reservation.materiel.nom}' n'est pas disponible pour cette période.")
                return redirect('liste_reservations_materiel')
            
            # Valider la réservation
            reservation.validee = True
            reservation.statut = 'validee'
            reservation.save()
            
            # Envoyer une notification au client
            from django.core.mail import send_mail
            subject = "Votre réservation de matériel a été validée"
            message = f"""
Bonjour {reservation.client.get_full_name() or reservation.client.username},

Votre réservation du matériel '{reservation.materiel.nom}' pour la période du {reservation.date_debut} au {reservation.date_fin} a été validée.

Détails de la réservation:
- Matériel: {reservation.materiel.nom}
- Période: du {reservation.date_debut} au {reservation.date_fin}
- Prix total: {reservation.prix_total} FCFA

Merci pour votre confiance.

Cordialement,
L'équipe de gestion agricole
            """
            
            try:
                send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [reservation.client.email])
            except:
                pass  # Ne pas échouer si l'email ne peut être envoyé
            
            messages.success(request, f"La réservation #{reservation.id} a été validée avec succès.")
            return redirect('liste_reservations_materiel')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la validation: {str(e)}")
    
    context = {'reservation': reservation}
    return render(request, 'admin/valider_reservation_materiel.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def annuler_reservation_materiel(request, reservation_id):
    """Annuler une réservation de matériel"""
    reservation = get_object_or_404(ReservationMateriel, id=reservation_id)
    
    if request.method == 'POST':
        commentaire_annulation = request.POST.get('commentaire_annulation', '')
        
        try:
            if not reservation.can_be_cancelled():
                messages.error(request, "Cette réservation ne peut plus être annulée.")
                return redirect('liste_reservations_materiel')
            
            # Annuler la réservation
            reservation.statut = 'annulee'
            reservation.validee = False
            reservation.commentaire_annulation = commentaire_annulation
            reservation.save()
            
            # Envoyer une notification au client
            from django.core.mail import send_mail
            subject = "Votre réservation de matériel a été annulée"
            message = f"""
Bonjour {reservation.client.get_full_name() or reservation.client.username},

Votre réservation du matériel '{reservation.materiel.nom}' pour la période du {reservation.date_debut} au {reservation.date_fin} a été annulée.

Motif: {commentaire_annulation or 'Aucun motif spécifié'}

N'hésitez pas à nous contacter pour plus d'informations.

Cordialement,
L'équipe de gestion agricole
            """
            
            try:
                send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [reservation.client.email])
            except:
                pass  # Ne pas échouer si l'email ne peut être envoyé
            
            messages.success(request, f"La réservation #{reservation.id} a été annulée avec succès.")
            return redirect('liste_reservations_materiel')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de l'annulation: {str(e)}")
    
    context = {'reservation': reservation}
    return render(request, 'admin/annuler_reservation_materiel.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def planning_locations(request):
    """Afficher le planning des locations avec calendrier"""
    from django.utils import timezone
    from datetime import datetime, timedelta
    import calendar
    
    # Récupérer le mois et l'année actuels ou ceux spécifiés
    today = timezone.now().date()
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))
    
    # Générer le calendrier
    cal = calendar.monthcalendar(year, month)
    month_name = calendar.month_name[month]
    
    # Récupérer les réservations pour ce mois
    reservations = ReservationMateriel.objects.filter(
        statut__in=['validee', 'en_cours'],
        date_debut__lte=timezone.now().date().replace(year=year, month=month, day=calendar.monthrange(year, month)[1]),
        date_fin__gte=timezone.now().date().replace(year=year, month=month, day=1)
    ).select_related('materiel', 'client')
    
    # Organiser les réservations par jour
    reservations_by_day = {}
    for reservation in reservations:
        current_date = reservation.date_debut
        while current_date <= reservation.date_fin:
            if current_date.month == month and current_date.year == year:
                day = current_date.day
                if day not in reservations_by_day:
                    reservations_by_day[day] = []
                reservations_by_day[day].append(reservation)
            current_date += timedelta(days=1)
    
    # Navigation entre mois
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    
    context = {
        'year': year,
        'month': month,
        'month_name': month_name,
        'calendar': cal,
        'reservations_by_day': reservations_by_day,
        'reservations': reservations,
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
        'materiels': Materiel.objects.all()
    }
    return render(request, 'admin/planning_locations.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def gerer_prix_materiel(request, materiel_id):
    """Gérer les prix d'un matériel"""
    materiel = get_object_or_404(Materiel, id=materiel_id)
    
    if request.method == 'POST':
        nouveau_prix = request.POST.get('prix_location_jour')
        
        try:
            ancien_prix = materiel.prix_location_jour
            materiel.prix_location_jour = nouveau_prix
            materiel.save()
            
            # Mettre à jour les réservations en attente avec le nouveau prix
            reservations_en_attente = materiel.reservations.filter(statut='en_attente')
            for reservation in reservations_en_attente:
                reservation.prix_jour_applique = nouveau_prix
                nb_jours = reservation.get_nombre_jours()
                reservation.prix_total = float(nouveau_prix) * nb_jours
                reservation.save()
            
            messages.success(request, 
                f"Le prix de location du matériel '{materiel.nom}' a été mis à jour de {ancien_prix} FCFA à {nouveau_prix} FCFA par jour. "
                f"{reservations_en_attente.count()} réservation(s) en attente ont été mise(s) à jour.")
            return redirect('gerant_materiels_agricoles')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la mise à jour du prix: {str(e)}")
    
    context = {
        'materiel': materiel,
        'reservations_en_attente': materiel.reservations.filter(statut='en_attente').count()
    }
    return render(request, 'admin/gerer_prix_materiel.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def statistiques_locations(request):
    """Afficher les statistiques des locations"""
    from django.db.models import Count, Sum, Avg
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    # Période sélectionnée
    periode = request.GET.get('periode', '30d')
    mapping = {'7d': 7, '30d': 30, '90d': 90, '365d': 365}
    days = mapping.get(periode, 30)
    
    since = timezone.now() - timedelta(days=days)
    
    # Statistiques générales
    total_reservations = ReservationMateriel.objects.count()
    reservations_validees = ReservationMateriel.objects.filter(validee=True).count()
    reservations_en_cours = ReservationMateriel.objects.filter(statut='en_cours').count()
    
    # Statistiques sur la période
    reservations_periode = ReservationMateriel.objects.filter(date_reservation__gte=since)
    chiffre_affaires_periode = ReservationMateriel.objects.filter(
        statut__in=['validee', 'en_cours', 'terminee'],
        date_reservation__gte=since
    ).aggregate(total=Sum('prix_total'))['total'] or 0
    
    # Top matériels les plus loués
    top_materiels = ReservationMateriel.objects.filter(
        statut__in=['validee', 'en_cours', 'terminee']
    ).values('materiel__nom').annotate(
        count=Count('id'),
        revenue=Sum('prix_total')
    ).order_by('-count')[:10]
    
    # Top clients
    top_clients = ReservationMateriel.objects.filter(
        statut__in=['validee', 'en_cours', 'terminee']
    ).values('client__username', 'client__first_name', 'client__last_name').annotate(
        count=Count('id'),
        total=Sum('prix_total')
    ).order_by('-count')[:10]
    
    # Évolution mensuelle
    evolution_mensuelle = []
    for i in range(6):
        month_start = timezone.now().replace(day=1) - timedelta(days=30*i)
        month_end = month_start.replace(day=calendar.monthrange(month_start.year, month_start.month)[1])
        
        stats_mois = ReservationMateriel.objects.filter(
            date_reservation__gte=month_start,
            date_reservation__lte=month_end
        ).aggregate(
            count=Count('id'),
            revenue=Sum('prix_total')
        )
        
        evolution_mensuelle.append({
            'month': month_start.strftime('%B %Y'),
            'count': stats_mois['count'] or 0,
            'revenue': stats_mois['revenue'] or 0
        })
    
    context = {
        'total_reservations': total_reservations,
        'reservations_validees': reservations_validees,
        'reservations_en_cours': reservations_en_cours,
        'reservations_periode': reservations_periode.count(),
        'chiffre_affaires_periode': chiffre_affaires_periode,
        'top_materiels': top_materiels,
        'top_clients': top_clients,
        'evolution_mensuelle': evolution_mensuelle,
        'periode': periode,
        'materiels': Materiel.objects.all()
    }
    return render(request, 'admin/statistiques_locations.html', context)
