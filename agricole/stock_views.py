# -*- coding: utf-8 -*-
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Sum, Q, F
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.utils import timezone
from django.views.decorators.http import require_POST
from .models import Produit, Stock, MouvementStock, Commande, DetailCommande
from .forms import StockForm, MouvementStockForm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.template.loader import get_template
import io

def is_gérant(user):
    """Vérifie si l'utilisateur est un gérant"""
    return user.is_authenticated and (user.is_staff or user.is_superuser or getattr(user, 'user_type', None) == '1')

@login_required
@user_passes_test(is_gérant)
def tableau_bord_stock(request):
    """
    Tableau de bord du stock avec statistiques et alertes
    """
    # Récupérer le gérant
    gerant = request.user
    
    # Statistiques générales
    total_produits = Produit.objects.count()
    total_stock = Stock.objects.filter(utilisateur=gerant).aggregate(
        total_quantite=Sum('quantite')
    )['total_quantite'] or 0
    
    # Produits en alerte
    stocks_alerte = Stock.objects.filter(
        utilisateur=gerant,
        quantite__lt=F('produit__seuil_alerte')
    ).select_related('produit')
    
    # Produits en rupture
    stocks_rupture = Stock.objects.filter(
        utilisateur=gerant,
        quantite__lte=0
    ).select_related('produit')
    
    # Mouvements récents
    mouvements_recents = MouvementStock.objects.filter(
        produit__stock__utilisateur=gerant
    ).select_related('produit', 'utilisateur').order_by('-date')[:10]
    
    # Valeur totale du stock
    valeur_stock = sum(
        stock.quantite * stock.produit.prix_unitaire 
        for stock in Stock.objects.filter(utilisateur=gerant).select_related('produit')
    )
    
    context = {
        'total_produits': total_produits,
        'total_stock': total_stock,
        'stocks_alerte': stocks_alerte,
        'stocks_rupture': stocks_rupture,
        'mouvements_recents': mouvements_recents,
        'valeur_stock': valeur_stock,
        'titre': 'Tableau de bord du stock',
    }
    
    return render(request, 'admin/stock/tableau_bord_stock.html', context)

@login_required
@user_passes_test(is_gérant)
def liste_stock_ameliore(request):
    """
    Liste améliorée du stock avec filtres et pagination
    """
    gerant = request.user
    
    # Récupérer les filtres
    search = request.GET.get('search', '')
    categorie_filtre = request.GET.get('categorie', '')
    statut_filtre = request.GET.get('statut', '')
    
    # Base query
    stocks = Stock.objects.filter(utilisateur=gerant).select_related('produit')
    
    # Appliquer les filtres
    if search:
        stocks = stocks.filter(produit__nom_produit__icontains=search)
    
    if categorie_filtre:
        stocks = stocks.filter(produit__categorie__id=categorie_filtre)
    
    # Filtrer par statut
    if statut_filtre == 'alerte':
        stocks = stocks.filter(quantite__lt=F('produit__seuil_alerte'))
    elif statut_filtre == 'rupture':
        stocks = stocks.filter(quantite__lte=0)
    elif statut_filtre == 'disponible':
        stocks = stocks.filter(quantite__gte=F('produit__seuil_alerte'))
    
    # Pagination
    paginator = Paginator(stocks, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calculer les statistiques pour chaque produit
    for stock in page_obj:
        stock.valeur_totale = stock.quantite * stock.produit.prix_unitaire
        stock.pourcentage_stock = (stock.quantite / stock.produit.seuil_alerte * 100) if stock.produit.seuil_alerte > 0 else 0
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'categorie_filtre': categorie_filtre,
        'statut_filtre': statut_filtre,
        'titre': 'Gestion du stock',
    }
    
    return render(request, 'admin/stock/liste_stock_ameliore.html', context)

@login_required
@user_passes_test(is_gérant)
def ajouter_entree_stock(request):
    """
    Ajouter une entrée en stock
    """
    if request.method == 'POST':
        produit_id = request.POST.get('produit')
        quantite = int(request.POST.get('quantite', 0))
        commentaire = request.POST.get('commentaire', '')
        
        if quantite <= 0:
            messages.error(request, 'La quantité doit être positive')
            return redirect('liste_stock_ameliore')
        
        produit = get_object_or_404(Produit, id=produit_id)
        stock = Stock.get_or_create_stock(request.user, produit)
        
        try:
            stock.augmenter_stock(quantite, commentaire or f"Entrée manuelle - {request.user.username}")
            messages.success(request, f'Entrée de {quantite} {produit.unite} de {produit.nom_produit} ajoutée avec succès')
        except ValueError as e:
            messages.error(request, str(e))
        
        return redirect('liste_stock_ameliore')
    
    # Récupérer les produits pour le formulaire
    produits = Produit.objects.all().order_by('nom_produit')
    
    context = {
        'produits': produits,
        'titre': 'Ajouter une entrée en stock',
    }
    
    return render(request, 'admin/stock/ajouter_entree_stock.html', context)

@login_required
@user_passes_test(is_gérant)
def valider_commande_avec_stock(request, commande_id):
    """
    Valider une commande en vérifiant et réduisant automatiquement le stock
    """
    commande = get_object_or_404(Commande, id=commande_id)
    
    # Assigner le gérant si non défini
    if not commande.gérant:
        commande.gérant = request.user
        commande.save()
    
    try:
        # Vérifier la disponibilité du stock
        disponible, message = commande.verifier_disponibilite_stock()
        
        if not disponible:
            messages.error(request, message)
            return redirect('liste_commandes')
        
        # Valider et réduire le stock
        commande.valider_et_reduire_stock()
        messages.success(request, f'Commande #{commande.id} validée et stock mis à jour avec succès')
        
    except ValueError as e:
        messages.error(request, str(e))
    
    return redirect('liste_commandes')

@login_required
@user_passes_test(is_gérant)
def annuler_commande_avec_stock(request, commande_id):
    """
    Annuler une commande et restituer automatiquement le stock
    """
    commande = get_object_or_404(Commande, id=commande_id)
    
    # Assigner le gérant si non défini
    if not commande.gérant:
        commande.gérant = request.user
        commande.save()
    
    try:
        commande.annuler_et_restituer_stock()
        messages.success(request, f'Commande #{commande.id} annulée et stock restitué avec succès')
        
    except ValueError as e:
        messages.error(request, str(e))
    
    return redirect('liste_commandes')

@login_required
@user_passes_test(is_gérant)
def details_commande_stock(request, commande_id):
    """
    Afficher les détails d'une commande avec le statut du stock
    """
    commande = get_object_or_404(Commande, id=commande_id)
    
    # Assigner le gérant si non défini
    if not commande.gérant:
        commande.gérant = request.user
        commande.save()
    
    # Récupérer le statut du stock pour chaque produit
    stock_status = commande.get_stock_status()
    
    # Vérifier si la commande peut être validée
    peut_valider = commande.statut == 'en_attente'
    if peut_valider:
        disponible, _ = commande.verifier_disponibilite_stock()
        peut_valider = disponible
    
    context = {
        'commande': commande,
        'stock_status': stock_status,
        'peut_valider': peut_valider,
        'titre': f'Détails de la commande #{commande.id}',
    }
    
    return render(request, 'admin/stock/details_commande_stock.html', context)

@login_required
@user_passes_test(is_gérant)
def mouvements_stock_ameliore(request):
    """
    Vue améliorée des mouvements de stock avec filtres
    """
    gerant = request.user
    
    # Récupérer les filtres
    produit_filtre = request.GET.get('produit', '')
    type_filtre = request.GET.get('type_mouvement', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    # Base query
    mouvements = MouvementStock.objects.filter(
        produit__stock__utilisateur=gerant
    ).select_related('produit', 'utilisateur')
    
    # Appliquer les filtres
    if produit_filtre:
        mouvements = mouvements.filter(produit_id=produit_filtre)
    
    if type_filtre:
        mouvements = mouvements.filter(type_mouvement=type_filtre)
    
    if date_debut:
        mouvements = mouvements.filter(date__gte=date_debut)
    
    if date_fin:
        mouvements = mouvements.filter(date__lte=date_fin)
    
    # Pagination
    paginator = Paginator(mouvements, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'produit_filtre': produit_filtre,
        'type_filtre': type_filtre,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'titre': 'Historique des mouvements de stock',
    }
    
    return render(request, 'admin/stock/mouvements_stock_ameliore.html', context)

@login_required
@user_passes_test(is_gérant)
def exporter_stock_ameliore_excel(request):
    """
    Exporter l'état du stock amélioré au format Excel
    """
    gerant = request.user
    
    # Récupérer les stocks
    stocks = Stock.objects.filter(utilisateur=gerant).select_related('produit')
    
    # Créer un classeur Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "État du stock"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajouter les en-têtes
    headers = ['Produit', 'Catégorie', 'Quantité', 'Unité', 'Seuil alerte', 'Prix unitaire', 'Valeur totale', 'Statut']
    for col_num, header in enumerate(headers, 1):
        cell = ws[f'{get_column_letter(col_num)}1']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Ajouter les données
    for row_num, stock in enumerate(stocks, 2):
        produit = stock.produit
        
        # Déterminer le statut
        if stock.est_en_rupture():
            statut = "Rupture"
        elif stock.est_en_alerte():
            statut = "Alerte"
        else:
            statut = "Normal"
        
        ws[f'A{row_num}'] = produit.nom_produit
        ws[f'B{row_num}'] = produit.categorie.nom if produit.categorie else ""
        ws[f'C{row_num}'] = stock.quantite
        ws[f'D{row_num}'] = produit.unite
        ws[f'E{row_num}'] = produit.seuil_alerte
        ws[f'F{row_num}'] = produit.prix_unitaire
        ws[f'G{row_num}'] = stock.quantite * produit.prix_unitaire
        ws[f'H{row_num}'] = statut
        
        # Colorer les lignes selon le statut
        if statut == "Rupture":
            for col in range(1, 9):
                ws[f'{get_column_letter(col)}{row_num}'].fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        elif statut == "Alerte":
            for col in range(1, 9):
                ws[f'{get_column_letter(col)}{row_num}'].fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    # Ajuster la largeur des colonnes
    column_widths = [30, 15, 10, 8, 12, 12, 12, 10]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Créer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=stock_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Sauvegarder le classeur dans la réponse
    wb.save(response)
    
    return response

@login_required
@user_passes_test(is_gérant)
@require_POST
def verifier_disponibilite_ajax(request):
    """
    Vérifier la disponibilité du stock via AJAX
    """
    produit_id = request.POST.get('produit_id')
    quantite = int(request.POST.get('quantite', 0))
    
    try:
        produit = Produit.objects.get(id=produit_id)
        stock = Stock.get_or_create_stock(request.user, produit)
        
        disponible = stock.peut_satisfaire(quantite)
        
        return JsonResponse({
            'disponible': disponible,
            'stock_actuel': stock.quantite,
            'message': 'Stock disponible' if disponible else f'Stock insuffisant. Disponible: {stock.quantite}, Demandé: {quantite}'
        })
        
    except Produit.DoesNotExist:
        return JsonResponse({'error': 'Produit non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)
